"""
Read Appendix (accessorial / ancillary tariffs) .xlsx from ./input and export
./output/<filename>.accessorial.json.

Layout (sheet "Europe english" or first sheet): two logical columns —
  Left:  A = CostName / section title, B = ApplyTo, C = CostPrice, D/E = minimum charge.
  Right: G = CostName / section title, H = ApplyTo, I = CostPrice.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _cell_nonempty(v: Any) -> bool:
    return bool(_cell_str(v))


def _normalize_currency_glyphs(s: str) -> str:
    return s.replace("\ufffd", "€")


def _is_price_only_basis(b_s: str, c: Any) -> bool:
    """When column C is empty, column B often carries the tariff outcome (Included, …)."""
    if _cell_nonempty(c):
        return False
    low = b_s.lower().strip()
    if not low:
        return False
    if low in (
        "included",
        "covered by asics",
        "spot pricing",
    ):
        return True
    if low.startswith("cf.") or low.startswith("outbound rates"):
        return True
    return False


def _split_apply_and_price(b: Any, c: Any) -> tuple[str | None, Any]:
    """Map Excel B/C into ApplyTo + CostPrice (B is not always 'apply')."""
    b_s = _normalize_currency_glyphs(_cell_str(b))
    if not b_s and not _cell_nonempty(c):
        return None, None
    if _is_price_only_basis(b_s, c):
        return None, _format_cost_price("", b_s) if b_s else None
    if not b_s and _cell_nonempty(c):
        return None, _format_cost_price("", c)
    apply_s = b_s
    price_out = _format_cost_price(apply_s, c) if _cell_nonempty(c) else None
    return (apply_s or None), price_out


def _format_cost_price(apply_to: str, value: Any) -> Any:
    """Prefer readable strings for small fractional rates; keep other numbers/strings."""
    if value is None:
        return None
    if isinstance(value, str):
        t = _normalize_currency_glyphs(value.strip())
        return t if t else None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:  # NaN
            return None
        a = apply_to.lower()
        if isinstance(value, float) and 0 < abs(value) < 1 and (
            "%" in apply_to or "percent" in a or "tariff" in a or "transport" in a
        ):
            pct = value * 100
            s = f"{pct:.6g}"
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            return f"{s}%"
        if value == 0 and isinstance(value, (int, float)):
            return "€0.00"
        if isinstance(value, float) and value == int(value):
            return int(value)
        return value
    return value


def _is_likely_section_header(
    title: str,
    b: Any,
    c: Any,
    d: Any,
    e: Any,
) -> bool:
    """Heuristic: standalone title row (no apply/price on same row)."""
    t = title.strip()
    if not t:
        return False
    if _cell_nonempty(b) or _cell_nonempty(c):
        return False
    if _cell_nonempty(d) or _cell_nonempty(e):
        return False
    if len(t) > 95 and not t.isupper():
        return False
    if len(t) > 140:
        return False
    return True


def _parse_tail_charge(
    d: Any, e: Any,
) -> tuple[str | None, str | None, Any | None]:
    """D/E column pair: minimum or maximum charge label + amount."""
    d_s = _cell_str(d).lower()
    if not _cell_nonempty(e):
        return None, None, None
    val = _format_cost_price("", e)
    lbl = _normalize_currency_glyphs(_cell_str(d))
    if "minimum" in d_s:
        return "min", lbl, val
    if "maximum" in d_s:
        return "max", lbl, val
    return None, None, None


def _parse_minimum_tail(
    d: Any, e: Any
) -> tuple[str | None, Any | None]:
    kind, lbl, val = _parse_tail_charge(d, e)
    if kind == "min":
        return lbl, val
    return None, None


class _SideParser:
    def __init__(self) -> None:
        self.blocks: list[dict[str, Any]] = []
        self._current_title: str | None = None

    def _ensure_block(self, title: str) -> None:
        title = _normalize_currency_glyphs(title.strip())
        if not title:
            return
        if self._current_title != title:
            self.blocks.append({"BlockTitle": title, "Items": []})
            self._current_title = title

    def _append_item(self, item: dict[str, Any]) -> None:
        if not self.blocks:
            self._ensure_block("UNCATEGORIZED")
        self.blocks[-1]["Items"].append(item)

    def flush_section_header(self, title: str) -> None:
        t = _normalize_currency_glyphs(_cell_str(title))
        if not t:
            return
        self._ensure_block(t)
        self._current_title = t

    def add_item_row(
        self,
        name: str,
        apply: Any,
        price: Any,
        d: Any,
        e: Any,
    ) -> None:
        n = _normalize_currency_glyphs(_cell_str(name))
        if not n:
            return
        if not self.blocks:
            self._ensure_block("UNCATEGORIZED")
        apply_s, price_out = _split_apply_and_price(apply, price)
        tail_kind, tail_lbl, tail_val = _parse_tail_charge(d, e)
        item: dict[str, Any] = {"CostName": n}
        if apply_s:
            item["ApplyTo"] = apply_s
        if price_out is not None:
            item["CostPrice"] = price_out
        if tail_kind == "min" and tail_val is not None:
            item["MinimumChargeLabel"] = tail_lbl
            item["MinimumCharge"] = tail_val
        elif tail_kind == "max" and tail_val is not None:
            item["MaximumChargeLabel"] = tail_lbl
            item["MaximumCharge"] = tail_val
        self._append_item(item)

    def merge_continuation_apply(self, apply: Any, price: Any, d: Any, e: Any) -> None:
        """Row with empty name: extend last item (e.g. regional split for same fee)."""
        if not self.blocks or not self.blocks[-1]["Items"]:
            return
        last = self.blocks[-1]["Items"][-1]
        apply_s, price_out = _split_apply_and_price(apply, price)
        tail_kind, tail_lbl, tail_val = _parse_tail_charge(d, e)
        if apply_s:
            prev = last.get("ApplyTo")
            last["ApplyTo"] = f"{prev}; {apply_s}" if prev else apply_s
        if price_out is not None:
            prev_p = last.get("CostPrice")
            if prev_p is not None:
                last["CostPrice"] = f"{prev_p}; {price_out}"
            else:
                last["CostPrice"] = price_out
        if tail_kind == "min" and tail_val is not None:
            last["MinimumChargeLabel"] = tail_lbl
            last["MinimumCharge"] = tail_val
        elif tail_kind == "max" and tail_val is not None:
            last["MaximumChargeLabel"] = tail_lbl
            last["MaximumCharge"] = tail_val


def _read_grid(ws: Any) -> list[list[Any]]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    max_col = max(max_col, 11)
    out: list[list[Any]] = []
    for ri in range(1, max_row + 1):
        row = [ws.cell(ri, ci).value for ci in range(1, max_col + 1)]
        out.append(row)
    return out


def _find_appendix_data_start(rows: list[list[Any]]) -> int:
    """Skip cover / title rows until the first appendix section row."""
    for i, row in enumerate(rows):
        blob = " ".join(_cell_str(c) for c in row[:11]).lower()
        if "administrative costs" in blob:
            return i
    return 0


def _parse_two_column_appendix(rows: list[list[Any]]) -> dict[str, Any]:
    start = _find_appendix_data_start(rows)
    rows = rows[start:]

    left = _SideParser()
    right = _SideParser()

    for row in rows:
        # pad to at least 11 cols
        while len(row) < 11:
            row.append(None)
        a, b, c = row[0], row[1], row[2]
        d, e = row[3], row[4]
        g, h, i = row[6], row[7], row[8]

        # --- Left column ---
        a_s, b_s, c_s = _cell_str(a), _cell_str(b), _cell_str(c)
        if a_s and _is_likely_section_header(a_s, b, c, d, e):
            left.flush_section_header(a_s)
        elif a_s and _cell_nonempty(b):
            left.add_item_row(a_s, b, c, d, e)
        elif a_s and not _cell_nonempty(b):
            # Name only (footnote / prose line)
            left.add_item_row(a_s, None, None, None, None)
        elif not a_s and _cell_nonempty(b):
            left.merge_continuation_apply(b, c, d, e)

        # --- Right column (G..I); min charge rarely used — same D,E indices wrong; use J,K if needed)
        j, k = row[9], row[10]
        g_s, h_s = _cell_str(g), _cell_str(h)
        if g_s and _is_likely_section_header(g_s, h, i, j, k):
            right.flush_section_header(g_s)
        elif g_s and _cell_nonempty(h):
            right.add_item_row(g_s, h, i, j, k)
        elif g_s and not _cell_nonempty(h):
            if _cell_nonempty(i):
                right.add_item_row(g_s, None, i, j, k)
            else:
                right.add_item_row(g_s, None, None, j, k)
        elif not g_s and _cell_nonempty(h):
            right.merge_continuation_apply(h, i, j, k)

    return {
        "LeftBlocks": left.blocks,
        "RightBlocks": right.blocks,
    }


def parse_accessorial_workbook(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb["Europe english"] if "Europe english" in wb.sheetnames else wb.active
        if ws is None:
            raise RuntimeError("Workbook has no active sheet")
        rows = _read_grid(ws)
        parsed = _parse_two_column_appendix(rows)
        return {
            "source_file": path.name,
            "sheet": ws.title,
            "AccessorialCostBlocks": {
                "Left": parsed["LeftBlocks"],
                "Right": parsed["RightBlocks"],
            },
        }
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert Appendix (accessorial) .xlsx to JSON.",
    )
    parser.add_argument(
        "input_xlsx",
        nargs="?",
        default=None,
        help="Path to .xlsx (default: pick from input/ interactively).",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output .json path (default: output/<stem>.accessorial.json).",
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=INPUT_DIR,
        help=f"Folder for interactive pick (default: {INPUT_DIR}).",
    )
    args = parser.parse_args()

    if args.input_xlsx:
        in_path = Path(args.input_xlsx)
        if not in_path.is_absolute() and not in_path.exists():
            cand = args.input_dir / in_path
            if cand.is_file():
                in_path = cand
    else:
        files = sorted(args.input_dir.glob("*.xlsx"), key=lambda p: p.name.lower())
        if not files:
            raise SystemExit(f"No .xlsx in {args.input_dir}")
        for i, p in enumerate(files, 1):
            print(f"  {i}. {p.name}")
        raw = input("Enter number: ").strip()
        if not raw.isdigit() or not (1 <= int(raw) <= len(files)):
            raise SystemExit("Invalid choice")
        in_path = files[int(raw) - 1]

    if not in_path.is_file():
        raise SystemExit(f"Input not found: {in_path}")

    out_path = (
        Path(args.output)
        if args.output
        else OUTPUT_DIR / f"{in_path.stem}.accessorial.json"
    )

    data = parse_accessorial_workbook(in_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
