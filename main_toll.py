"""
Extract toll tabs (sheet name contains 'Toll', case-insensitive) from the
Europe transport tariff .xlsx into JSON (e.g.
``Copy of Tariffs Europe for DCS - 2024-2027 - 20240401 - V14 (2).xlsx``).
Toll is not in the appendix workbook.

Output: cost matrices (zone × weight bands), zoning rows (postal CP layout
and/or all-caps ZONE + country columns for transit), and CostNameApplicable
links to domestic vs transit cost tables.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"

MAX_COLS = 40
FLAT_RATE_SNIPPET = "flat rate pricing expressed in euros"


def _discover_toll_workbooks(directory: Path) -> list[Path]:
    """Europe transport tariff workbooks (toll tabs); exclude appendix files."""
    all_xlsx = sorted(directory.glob("*.xlsx"), key=lambda p: p.name.lower())
    europe = [
        p
        for p in all_xlsx
        if "europe" in p.name.lower() and "appendix" not in p.name.lower()
    ]
    if europe:
        return europe
    return [p for p in all_xlsx if "appendix" not in p.name.lower()]


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _cell_nonempty(v: Any) -> bool:
    return bool(_cell_str(v))


def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def _read_sheet_grid(ws: Any, max_row: int | None = None) -> list[list[Any]]:
    mr = max_row or (ws.max_row or 0)
    mc = min(MAX_COLS, ws.max_column or 0)
    mc = max(mc, 18)
    out: list[list[Any]] = []
    for ri in range(1, mr + 1):
        out.append([ws.cell(ri, ci).value for ci in range(1, mc + 1)])
    return out


def _is_weight_header_cell(s: str) -> bool:
    """Kg band columns plus non-kg cap bands (Full truck / Complet) on the same header row."""
    low = _norm_ws(s).lower()
    if not low:
        return False
    if "from" in low and "kg" in low:
        return True
    if "full truck" in low:
        return True
    if low in ("complet", "complete", "ftl"):
        return True
    return False


def _weight_header_score(row: list[Any]) -> int:
    return sum(1 for c in row if _is_weight_header_cell(_cell_str(c)))


def _is_zone_label(s: str) -> bool:
    t = _norm_ws(_cell_str(s))
    if not t:
        return False
    return bool(re.match(r"(?i)zone\s*\d+\s*:", t)) or bool(re.match(r"(?i)^zone\s*\d+\s*$", t))


def _looks_like_toll_title(s: str) -> bool:
    low = s.lower()
    if "toll" not in low and "maut" not in low:
        return False
    if len(_norm_ws(s)) < 12:
        return False
    return True


def _find_toll_title_in_row(row: list[Any]) -> str | None:
    best = ""
    for c in row:
        t = _norm_ws(_cell_str(c))
        if _looks_like_toll_title(t) and len(t) > len(best):
            best = t
    return best or None


def _is_flat_rate_footer(row: list[Any]) -> bool:
    blob = " ".join(_cell_str(c).lower() for c in row[:20])
    return FLAT_RATE_SNIPPET in blob


def _is_germany_domestic_zoning_row(row: list[Any]) -> bool:
    """Row 38 style: Zone 1 :, Germany, CP 66."""
    if len(row) < 3:
        return False
    a, b = _cell_str(row[0]), _cell_str(row[1])
    if not _is_zone_label(a):
        return False
    return "germany" in b.lower()


def _is_austria_cp_zoning_row(row: list[Any]) -> bool:
    """Zone title in A, postal code list in B (not a numeric toll cell)."""
    if len(row) < 2:
        return False
    a, b = row[0], row[1]
    if not _is_zone_label(_cell_str(a)):
        return False
    if isinstance(b, (int, float)):
        return False
    bs = _cell_str(b).lower()
    if not bs or "germany" in bs:
        return False
    if bs.startswith("cp") or bs.startswith("cp "):
        return True
    if " cp " in bs or "cp " in bs:
        return True
    if " to " in bs and re.search(r"\d", bs) and len(bs) > 8:
        return True
    return False


def _find_zoning_start_index(rows: list[list[Any]], tab_country_hint: str) -> int | None:
    for i, row in enumerate(rows):
        if _is_germany_domestic_zoning_row(row):
            return i
        if tab_country_hint == "austria" and _is_austria_cp_zoning_row(row):
            return i
    return None


def _parse_weight_columns(header_row: list[Any]) -> list[tuple[int, str]]:
    out: list[tuple[int, str]] = []
    for j, c in enumerate(header_row):
        s = _norm_ws(_cell_str(c))
        if _is_weight_header_cell(s):
            out.append((j, s))
    return out


def _parse_zone_data_row(
    row: list[Any],
    weight_cols: list[tuple[int, str]],
) -> dict[str, Any] | None:
    zone_j = None
    zone_text = ""
    for j, c in enumerate(row):
        if _is_zone_label(_cell_str(c)):
            zone_j = j
            zone_text = _norm_ws(_cell_str(c))
            break
    if zone_j is None:
        return None
    bands: list[dict[str, Any]] = []
    for wj, wlabel in weight_cols:
        if wj >= len(row):
            continue
        if wj <= zone_j:
            continue
        v = row[wj]
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if isinstance(v, (int, float)):
            bands.append({"Weight": wlabel, "Cost": float(v) if isinstance(v, float) else v})
        else:
            bands.append({"Weight": wlabel, "Cost": _norm_ws(_cell_str(v))})
    if not bands:
        return None
    return {"ZoneName": zone_text, "Bands": bands}


def _row_has_numeric_weight_bands(
    row: list[Any],
    weight_cols: list[tuple[int, str]],
    min_count: int = 2,
) -> bool:
    n = 0
    for wj, _ in weight_cols:
        if wj < len(row) and isinstance(row[wj], (int, float)):
            n += 1
    return n >= min_count


def _parse_implicit_zone_rate_row(
    row: list[Any],
    weight_cols: list[tuple[int, str]],
    zone_name: str,
) -> dict[str, Any] | None:
    """Single tariff row with weights in columns but no `Zone N` label (e.g. Austria transit)."""
    bands: list[dict[str, Any]] = []
    for wj, wlabel in weight_cols:
        if wj >= len(row):
            continue
        v = row[wj]
        if v is None or (isinstance(v, str) and not str(v).strip()):
            continue
        if isinstance(v, (int, float)):
            bands.append({"Weight": wlabel, "Cost": float(v) if isinstance(v, float) else v})
        else:
            bands.append({"Weight": wlabel, "Cost": _norm_ws(_cell_str(v))})
    if len(bands) < 2:
        return None
    return {"ZoneName": zone_name, "Bands": bands}


def _parse_rate_row_for_toll_table(
    row: list[Any],
    weight_cols: list[tuple[int, str]],
    *,
    implicit_zone_name: str | None = None,
) -> dict[str, Any] | None:
    zd = _parse_zone_data_row(row, weight_cols)
    if zd is not None:
        return zd
    if implicit_zone_name and _row_has_numeric_weight_bands(row, weight_cols):
        return _parse_implicit_zone_rate_row(row, weight_cols, implicit_zone_name)
    return None


def _zoning_triples_from_row(row: list[Any]) -> list[tuple[str, str, str]]:
    """Each (ZoneName, Country, PostalCode) from columns after each Zone label."""
    out: list[tuple[str, str, str]] = []
    for j, _ in enumerate(row):
        zn = _norm_ws(_cell_str(row[j]))
        if not zn or not _is_zone_label(zn):
            continue
        c1 = row[j + 1] if j + 1 < len(row) else None
        c2 = row[j + 2] if j + 2 < len(row) else None
        co = _norm_ws(_cell_str(c1))
        pc = _norm_ws(_cell_str(c2))
        if not co and not pc:
            continue
        out.append((zn, co, pc))
    return out


def _parse_zoning_pairs_germany(rows: list[list[Any]], start: int, end: int) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for i in range(start, min(end, len(rows))):
        row = rows[i]
        while len(row) < 12:
            row.append(None)
        for zn, co, pc in _zoning_triples_from_row(row):
            out.append(
                {
                    "ZoneName": zn,
                    "Country": co or None,
                    "PostalCode": pc or None,
                }
            )
    return out


def _parse_zoning_austria(rows: list[list[Any]], start: int, end: int) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for i in range(start, min(end, len(rows))):
        row = rows[i]
        a, b = _cell_str(row[0]), _cell_str(row[1]) if len(row) > 1 else ""
        if not _is_zone_label(a):
            continue
        if not _norm_ws(b):
            continue
        out.append(
            {
                "ZoneName": _norm_ws(a),
                "Country": "Austria",
                "PostalCode": _norm_ws(b),
            }
        )
    return out


def _infer_tab_country_hint(tab_name: str) -> str:
    low = tab_name.lower()
    if "austria" in low:
        return "austria"
    if "germany" in low:
        return "germany"
    return ""


def _next_weight_header_row_index(
    rows: list[list[Any]], title_row: int, limit: int, max_gap: int = 6
) -> int | None:
    for d in range(1, max_gap + 1):
        j = title_row + d
        if j >= limit:
            return None
        if _weight_header_score(rows[j]) >= 2:
            return j
    return None


def _parse_cost_tables(
    rows: list[list[Any]],
    zoning_start: int | None,
    tab_hint: str,
) -> tuple[list[dict[str, Any]], int | None]:
    """
    Walk rows and build CostTables until zoning or sheet end.
    Returns (tables, index_after_domestic_block) for optional transit continuation.
    """
    tables: list[dict[str, Any]] = []
    i = 0
    limit = zoning_start if zoning_start is not None else len(rows)
    last_zone_label: str | None = None

    while i < limit:
        row = rows[i]
        title = _find_toll_title_in_row(row)
        if title:
            wh = _next_weight_header_row_index(rows, i, limit)
            if wh is not None:
                weight_cols = _parse_weight_columns(rows[wh])
                i = wh + 1
                rates: list[dict[str, Any]] = []
                matrix_seg = 1
                while i < limit:
                    r = rows[i]
                    if _weight_header_score(r) >= 2:
                        weight_cols = _parse_weight_columns(r)
                        i += 1
                        last_zone_label = None
                        matrix_seg += 1
                        continue
                    if _is_flat_rate_footer(r):
                        i += 1
                        break
                    if _find_toll_title_in_row(r) and _weight_header_score(r) < 2:
                        break
                    if zoning_start is not None and i >= zoning_start:
                        break
                    if tab_hint == "germany" and _is_germany_domestic_zoning_row(r):
                        break
                    if tab_hint == "austria" and _is_austria_cp_zoning_row(r):
                        break
                    zd = _parse_rate_row_for_toll_table(r, weight_cols, implicit_zone_name=None)
                    if zd:
                        last_zone_label = zd["ZoneName"]
                        if matrix_seg > 1:
                            zd["WeightMatrixSegment"] = matrix_seg
                        rates.append(zd)
                        i += 1
                        continue
                    if last_zone_label and weight_cols:
                        nums_only = all(
                            wj < len(r) and isinstance(r[wj], (int, float))
                            for wj, _ in weight_cols
                        )
                        if nums_only:
                            synthetic = {
                                "ZoneName": last_zone_label,
                                "Bands": [],
                            }
                            if matrix_seg > 1:
                                synthetic["WeightMatrixSegment"] = matrix_seg
                            for wj, wlabel in weight_cols:
                                if wj < len(r) and isinstance(r[wj], (int, float)):
                                    synthetic["Bands"].append(
                                        {"Weight": wlabel, "Cost": float(r[wj])}
                                    )
                            if synthetic["Bands"]:
                                rates.append(synthetic)
                            i += 1
                            continue
                    i += 1
                tables.append({"CostName": title, "Rates": rates})
                continue
        i += 1

    return tables, None


def _parse_transit_tail(
    rows: list[list[Any]],
    start: int,
    tab_hint: str,
) -> list[dict[str, Any]]:
    """Parse additional cost tables after zoning (e.g. TRANSIT ON GERMANY)."""
    tables: list[dict[str, Any]] = []
    i = start
    last_zone_label: str | None = None
    while i < len(rows):
        row = rows[i]
        if _cell_str(row[0]).lower().startswith("xxx"):
            break
        title = _find_toll_title_in_row(row)
        if title:
            wh = _next_weight_header_row_index(rows, i, len(rows))
            if wh is not None:
                weight_cols = _parse_weight_columns(rows[wh])
                i = wh + 1
                rates: list[dict[str, Any]] = []
                matrix_seg = 1
                while i < len(rows):
                    r = rows[i]
                    if _cell_str(r[0]).lower().startswith("xxx"):
                        break
                    if _weight_header_score(r) >= 2:
                        weight_cols = _parse_weight_columns(r)
                        last_zone_label = None
                        i += 1
                        matrix_seg += 1
                        continue
                    if _is_flat_rate_footer(r):
                        i += 1
                        break
                    if _find_toll_title_in_row(r) and _weight_header_score(r) < 2:
                        break
                    zd = _parse_rate_row_for_toll_table(
                        r,
                        weight_cols,
                        implicit_zone_name="All distances",
                    )
                    if zd:
                        last_zone_label = zd["ZoneName"]
                        if matrix_seg > 1:
                            zd["WeightMatrixSegment"] = matrix_seg
                        rates.append(zd)
                        i += 1
                        continue
                    if last_zone_label and weight_cols:
                        if all(
                            wj < len(r) and isinstance(r[wj], (int, float))
                            for wj, _ in weight_cols
                        ):
                            syn = {"ZoneName": last_zone_label, "Bands": []}
                            if matrix_seg > 1:
                                syn["WeightMatrixSegment"] = matrix_seg
                            for wj, wlabel in weight_cols:
                                if wj < len(r) and isinstance(r[wj], (int, float)):
                                    syn["Bands"].append(
                                        {"Weight": wlabel, "Cost": float(r[wj])}
                                    )
                            if syn["Bands"]:
                                rates.append(syn)
                            i += 1
                            continue
                    if _find_toll_title_in_row(r):
                        break
                    i += 1
                tables.append({"CostName": title, "Rates": rates})
                continue
        i += 1
    return tables


def _find_zoning_end_germany(rows: list[list[Any]], start: int) -> int:
    for i in range(start + 1, len(rows)):
        if _find_toll_title_in_row(rows[i]):
            return i
    return len(rows)


def _find_zoning_end_austria(rows: list[list[Any]], start: int) -> int:
    for i in range(start + 1, len(rows)):
        r = rows[i]
        if _find_toll_title_in_row(r) and "transit" in _find_toll_title_in_row(r).lower():
            return i
        if _weight_header_score(r) >= 2 and i > start + 1:
            return i
    return len(rows)


def _is_strict_zone_matrix_label(s: str) -> bool:
    """All-caps style `ZONE 1` (no colon) — country-matrix block, not `Zone 1 :` CP rows."""
    return bool(re.fullmatch(r"(?i)zone\s*\d+", _norm_ws(s)))


def _looks_like_country_name_token(s: str) -> bool:
    t = _norm_ws(s)
    if len(t) < 3:
        return False
    low = t.lower()
    if low in {"none", "xxx"}:
        return False
    if "concerned" in low:
        return False
    if _is_weight_header_cell(t):
        return False
    if low.startswith("flat rate"):
        return False
    if re.fullmatch(r"[\d.,\s€%-]+", t):
        return False
    if not re.search(r"[a-zA-Z]", t):
        return False
    return True


def _is_country_matrix_seed_row(row: list[Any]) -> bool:
    """`ZONE 1` + country names in following columns (not `ZONE 1` + toll amounts)."""
    if len(row) < 2:
        return False
    a = _cell_str(row[0])
    if not _is_strict_zone_matrix_label(a):
        return False
    b = row[1]
    if isinstance(b, (int, float)):
        return False
    bs = _cell_str(b)
    return _looks_like_country_name_token(bs)


def _find_first_country_matrix_zoning_row(rows: list[list[Any]]) -> int | None:
    for i, row in enumerate(rows):
        if _is_country_matrix_seed_row(row):
            return i
    return None


def _finalize_country_column_zone(
    cur: dict[str, Any],
    transit_cost_names: list[str],
) -> dict[str, Any]:
    countries: list[str] = list(cur.get("Countries") or [])
    return {
        "ZoneName": cur["ZoneName"],
        "Country": ", ".join(countries),
        "Countries": countries,
        "ZoningFormat": "country_columns",
        "CostNameApplicable": list(transit_cost_names),
    }


def _parse_country_column_zoning(
    rows: list[list[Any]],
    start_i: int,
    transit_cost_names: list[str],
) -> list[dict[str, Any]]:
    """
    Blocks like `ZONE 1 | ALBANIA | BULGARIA | ...` with optional wrap rows
    (empty first column, more countries). Applies to transit toll tables only.
    """
    out: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    i = start_i
    while i < len(rows):
        row = rows[i]
        if _cell_str(row[0]).lower().startswith("xxx"):
            if current:
                out.append(_finalize_country_column_zone(current, transit_cost_names))
                current = None
            break
        if _is_country_matrix_seed_row(row):
            if current:
                out.append(_finalize_country_column_zone(current, transit_cost_names))
            zn = _norm_ws(_cell_str(row[0])).upper()
            countries: list[str] = []
            for c in row[1:]:
                s = _norm_ws(_cell_str(c))
                if s and _looks_like_country_name_token(s):
                    countries.append(s)
            current = {"ZoneName": zn, "Countries": countries}
            i += 1
            continue
        if current and not _cell_str(row[0]):
            added = False
            for c in row[1:]:
                s = _norm_ws(_cell_str(c))
                if s and _looks_like_country_name_token(s):
                    current["Countries"].append(s)
                    added = True
            if added:
                i += 1
                continue
            if not any(_cell_nonempty(c) for c in row[1:]):
                i += 1
                continue
        if current:
            out.append(_finalize_country_column_zone(current, transit_cost_names))
            current = None
        if out:
            break
        i += 1
    if current:
        out.append(_finalize_country_column_zone(current, transit_cost_names))
    return out


def _parse_austria_transit_concerned_zoning(
    rows: list[list[Any]],
    transit_cost_names: list[str],
) -> list[dict[str, Any]]:
    """
    Rows under 'Concerned countries' listing destination states for Austrian transit toll.
    """
    if not transit_cost_names:
        return []
    for i, row in enumerate(rows):
        a = _cell_str(row[0]).lower()
        if "concerned" not in a or "countr" not in a:
            continue
        countries: list[str] = []
        for c in row[1:]:
            s = _norm_ws(_cell_str(c))
            if s and _looks_like_country_name_token(s):
                countries.append(s)
        if i + 1 < len(rows):
            r2 = rows[i + 1]
            if not _cell_str(r2[0]):
                for c in r2[1:]:
                    s = _norm_ws(_cell_str(c))
                    if s and _looks_like_country_name_token(s):
                        countries.append(s)
        if not countries:
            return []
        return [
            {
                "ZoneName": "Concerned countries",
                "Country": ", ".join(countries),
                "Countries": countries,
                "ZoningFormat": "transit_concerned_countries",
                "CostNameApplicable": list(transit_cost_names),
            }
        ]
    return []


def parse_toll_sheet(ws: Any) -> dict[str, Any]:
    rows = _read_sheet_grid(ws)
    tab = ws.title or "Toll"
    hint = _infer_tab_country_hint(tab)
    zstart = _find_zoning_start_index(rows, hint)

    domestic_tables, _ = _parse_cost_tables(rows, zstart, hint)

    zoning: list[dict[str, Any]] = []
    transit_tables: list[dict[str, Any]] = []
    applicable = [t["CostName"] for t in domestic_tables]

    if zstart is not None:
        if hint == "germany":
            zend = _find_zoning_end_germany(rows, zstart)
            raw = _parse_zoning_pairs_germany(rows, zstart, zend)
            for z in raw:
                z["CostNameApplicable"] = list(applicable)
            zoning = raw
            transit_tables = _parse_transit_tail(rows, zend, hint)
        elif hint == "austria":
            zend = _find_zoning_end_austria(rows, zstart)
            raw = _parse_zoning_austria(rows, zstart, zend)
            for z in raw:
                z["CostNameApplicable"] = list(applicable)
            zoning = raw
            transit_tables = _parse_transit_tail(rows, zend, hint)
        else:
            raw = _parse_zoning_pairs_germany(rows, zstart, len(rows))
            for z in raw:
                z["CostNameApplicable"] = list(applicable)
            zoning = raw

    all_tables = domestic_tables + transit_tables

    transit_names = [
        t["CostName"]
        for t in transit_tables
        if "transit" in (t.get("CostName") or "").lower()
    ]
    if not transit_names:
        transit_names = [
            t["CostName"]
            for t in all_tables
            if "transit" in (t.get("CostName") or "").lower()
        ]

    if hint == "austria" and transit_names:
        zoning.extend(_parse_austria_transit_concerned_zoning(rows, transit_names))

    cm_row = _find_first_country_matrix_zoning_row(rows)
    if cm_row is not None:
        zoning.extend(_parse_country_column_zoning(rows, cm_row, transit_names))

    return {
        "Tab": tab,
        "CostTables": all_tables,
        "Zoning": zoning,
    }


def _is_czech_lkw_tariff_title(title: str | None) -> bool:
    if not title:
        return False
    low = title.lower()
    if "czech" not in low:
        return False
    return "lkw" in low or "maut" in low or "<<" in title


def _normalize_czech_lkw_row(row: list[Any], header: bool) -> list[Any]:
    """Map B=zone / D+=weights layout to the same shape as DE/AT (zone col A, weights from C)."""
    if len(row) < 4:
        return list(row)
    tail = list(row[3:])
    if header:
        return [None, None] + tail
    return [row[1], None] + tail


def _is_czech_domestic_zoning_row(row: list[Any]) -> bool:
    if len(row) < 5:
        return False
    if not _is_zone_label(_cell_str(row[1])):
        return False
    return "czech" in _cell_str(row[2]).lower()


def _is_czech_transit_dest_zoning_row(row: list[Any]) -> bool:
    if len(row) < 4:
        return False
    if not _is_zone_label(_cell_str(row[1])):
        return False
    if "czech" in _cell_str(row[2]).lower():
        return False
    return any(
        _looks_like_country_name_token(_cell_str(c)) for c in row[3:15]
    )


def _find_xxx_row_index(rows: list[list[Any]]) -> int:
    for i, row in enumerate(rows):
        if _cell_str(row[0]).lower().startswith("xxx"):
            return i
    return len(rows)


def _find_czech_domestic_cost_end(
    rows: list[list[Any]], wh: int, limit: int
) -> int:
    for i in range(wh + 1, limit):
        if _is_flat_rate_footer(rows[i]):
            return i
        if _is_czech_domestic_zoning_row(rows[i]):
            return i
    return limit


def _find_czech_transit_cost_end(
    rows: list[list[Any]], wh: int, limit: int
) -> int:
    for i in range(wh + 1, limit):
        if _is_czech_transit_dest_zoning_row(rows[i]):
            return i
        if _is_flat_rate_footer(rows[i]):
            return i
    return limit


def _find_czech_postal_zoning_start(
    rows: list[list[Any]], search_from: int, limit: int
) -> int | None:
    for i in range(search_from, limit):
        if _is_czech_domestic_zoning_row(rows[i]):
            return i
    return None


def _parse_czech_lkw_matrix(
    rows: list[list[Any]],
    title_row_idx: int,
    cost_end_exclusive: int,
) -> dict[str, Any]:
    title = _find_toll_title_in_row(rows[title_row_idx]) or "Czech toll"
    wh = _next_weight_header_row_index(rows, title_row_idx, cost_end_exclusive)
    if wh is None or wh >= cost_end_exclusive:
        return {"CostName": title, "Rates": []}
    weight_cols = _parse_weight_columns(
        _normalize_czech_lkw_row(rows[wh], True)
    )
    i = wh + 1
    rates: list[dict[str, Any]] = []
    matrix_seg = 1
    last_zone_label: str | None = None
    while i < cost_end_exclusive:
        r = rows[i]
        nhr = _normalize_czech_lkw_row(r, True)
        if _weight_header_score(nhr) >= 2:
            weight_cols = _parse_weight_columns(nhr)
            last_zone_label = None
            matrix_seg += 1
            i += 1
            continue
        if _is_flat_rate_footer(r):
            i += 1
            break
        rn = _normalize_czech_lkw_row(r, False)
        zd = _parse_rate_row_for_toll_table(rn, weight_cols, implicit_zone_name=None)
        if zd:
            last_zone_label = zd["ZoneName"]
            if matrix_seg > 1:
                zd["WeightMatrixSegment"] = matrix_seg
            rates.append(zd)
            i += 1
            continue
        if last_zone_label and weight_cols:
            if all(
                wj < len(rn) and isinstance(rn[wj], (int, float))
                for wj, _ in weight_cols
            ):
                syn: dict[str, Any] = {"ZoneName": last_zone_label, "Bands": []}
                if matrix_seg > 1:
                    syn["WeightMatrixSegment"] = matrix_seg
                for wj, wlabel in weight_cols:
                    if wj < len(rn) and isinstance(rn[wj], (int, float)):
                        syn["Bands"].append(
                            {"Weight": wlabel, "Cost": float(rn[wj])}
                        )
                if syn["Bands"]:
                    rates.append(syn)
                i += 1
                continue
        i += 1
    return {"CostName": title, "Rates": rates}


def _collect_czech_domestic_zoning(
    rows: list[list[Any]], z0: int, limit: int, applicable: list[str]
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    i = z0
    while i < limit:
        row = rows[i]
        if not _is_czech_domestic_zoning_row(row):
            break
        zn = _norm_ws(_cell_str(row[1]))
        co = _norm_ws(_cell_str(row[2]))
        pc = _norm_ws(_cell_str(row[4])) if len(row) > 4 else ""
        out.append(
            {
                "ZoneName": zn,
                "Country": co or None,
                "PostalCode": pc or None,
                "CostNameApplicable": list(applicable),
            }
        )
        i += 1
    return out


def _collect_czech_transit_dest_zoning(
    rows: list[list[Any]], z0: int, limit: int, applicable: list[str]
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    i = z0
    while i < limit:
        row = rows[i]
        if not _is_czech_transit_dest_zoning_row(row):
            break
        countries: list[str] = []
        for c in row[3:]:
            s = _norm_ws(_cell_str(c))
            if s and _looks_like_country_name_token(s):
                countries.append(s)
        zn = _norm_ws(_cell_str(row[1]))
        out.append(
            {
                "ZoneName": zn,
                "Country": ", ".join(countries),
                "Countries": countries,
                "ZoningFormat": "transit_destination_zones",
                "CostNameApplicable": list(applicable),
            }
        )
        i += 1
    return out


def parse_other_countries_toll_sheet(ws: Any) -> dict[str, Any]:
    """
    TOLL ON OTHER COUNTRIES: structured Czech LKW MAUT blocks (same CostTables /
    Zoning shape as Germany/Austria); remaining surcharge lines stay in
    UnstructuredBlocks.
    """
    rows = _read_sheet_grid(ws, max_row=min(ws.max_row or 0, 250))
    xxx = _find_xxx_row_index(rows)
    czech_title_rows = [
        i
        for i in range(len(rows))
        if (t := _find_toll_title_in_row(rows[i])) and _is_czech_lkw_tariff_title(t)
    ]
    czech_spans: set[int] = set()
    for idx, ti in enumerate(czech_title_rows):
        nxt = czech_title_rows[idx + 1] if idx + 1 < len(czech_title_rows) else xxx
        for r in range(ti, min(nxt, len(rows))):
            czech_spans.add(r)

    cost_tables: list[dict[str, Any]] = []
    zoning: list[dict[str, Any]] = []

    for idx, ti in enumerate(czech_title_rows):
        nxt = czech_title_rows[idx + 1] if idx + 1 < len(czech_title_rows) else xxx
        title = _find_toll_title_in_row(rows[ti]) or ""
        is_transit = "transit" in title.lower()
        wh = _next_weight_header_row_index(rows, ti, nxt)
        if wh is None:
            continue
        cost_end = (
            _find_czech_transit_cost_end(rows, wh, nxt)
            if is_transit
            else _find_czech_domestic_cost_end(rows, wh, nxt)
        )
        ct = _parse_czech_lkw_matrix(rows, ti, cost_end)
        if ct.get("Rates"):
            cost_tables.append(ct)
        if is_transit:
            zoning.extend(
                _collect_czech_transit_dest_zoning(rows, cost_end, nxt, [title])
            )
        else:
            flat_i = None
            for j in range(wh + 1, nxt):
                if _is_flat_rate_footer(rows[j]):
                    flat_i = j
                    break
            if flat_i is not None:
                z0 = _find_czech_postal_zoning_start(rows, flat_i + 1, nxt)
                if z0 is not None:
                    zoning.extend(
                        _collect_czech_domestic_zoning(rows, z0, nxt, [title])
                    )

    blocks: list[dict[str, Any]] = []
    cur: dict[str, Any] | None = None
    for ri, row in enumerate(rows):
        if ri in czech_spans:
            continue
        line = _norm_ws(" | ".join(_cell_str(c) for c in row[:20] if _cell_str(c)))
        if not line:
            continue
        low = line.lower()
        if "toll on" in low and len(line) < 80:
            cur = {"SectionTitle": line, "Rows": []}
            blocks.append(cur)
            continue
        if cur is not None:
            cur["Rows"].append(line)
    return {
        "Tab": ws.title or "Toll",
        "CostTables": cost_tables,
        "Zoning": zoning,
        "UnstructuredBlocks": blocks,
    }


def extract_workbook_tolls(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        toll_sheets: list[dict[str, Any]] = []
        for name in wb.sheetnames:
            if "toll" not in name.lower():
                continue
            ws = wb[name]
            if "other countries" in name.lower():
                toll_sheets.append(parse_other_countries_toll_sheet(ws))
            else:
                toll_sheets.append(parse_toll_sheet(ws))
        return {"source_file": path.name, "TollSheets": toll_sheets}
    finally:
        wb.close()


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Export Toll* tabs from the Europe transport tariff .xlsx to JSON "
            "(<stem>.toll.json)."
        ),
    )
    ap.add_argument(
        "input_xlsx",
        nargs="?",
        default=None,
        help="Europe tariffs .xlsx (toll tabs). Default: pick from input/.",
    )
    ap.add_argument("-o", "--output", default=None)
    ap.add_argument("--input-dir", type=Path, default=INPUT_DIR)
    args = ap.parse_args()

    if args.input_xlsx:
        in_path = Path(args.input_xlsx)
        if not in_path.is_absolute() and not in_path.exists():
            cand = args.input_dir / in_path
            if cand.is_file():
                in_path = cand
    else:
        files = _discover_toll_workbooks(args.input_dir)
        if not files:
            raise SystemExit(f"No suitable .xlsx in {args.input_dir}")
        if len(files) == 1:
            in_path = files[0]
            print(f"Using Europe tariffs workbook: {in_path.name}")
        else:
            print("Europe tariff workbooks with toll tabs (not appendix):")
            for i, p in enumerate(files, 1):
                print(f"  {i}. {p.name}")
            raw = input("Enter number: ").strip()
            if not raw.isdigit() or not (1 <= int(raw) <= len(files)):
                raise SystemExit("Invalid choice")
            in_path = files[int(raw) - 1]

    if not in_path.is_file():
        raise SystemExit(f"Input not found: {in_path}")

    out = Path(args.output) if args.output else OUTPUT_DIR / f"{in_path.stem}.toll.json"
    data = extract_workbook_tolls(in_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Wrote: {out}")


if __name__ == "__main__":
    main()
