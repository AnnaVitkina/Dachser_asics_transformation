"""
End-to-end tariff pipeline.

Folder paths are fixed below (INPUT_DIR, PROCESSING_DIR, OUTPUT_DIR).
You choose which .xlsx is transport costs and which is accessorial.
Run: python pipeline.py
"""

from __future__ import annotations

import json
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook

import main as transport_main
import main_accessorial
import main_toll
from accessorial_to_xlsx import (
    build_ra_table,
    write_report,
    write_xlsx,
)
from excel_layout import (
    apply_accessorial_sheet_layout,
    apply_toll_pct_sheet_layout,
    apply_toll_structured_sheet_layout,
    copy_worksheet,
)
from toll_to_excel import (
    transform_toll_json_to_workbook,
    write_country_regions_catalog,
    write_postal_zones_catalog,
)
from transformation_to_xlsx import transform_json_to_xlsx

# --- Hardcoded folder paths (edit here if needed) ---
INPUT_DIR = Path(r"C:\Users\avitkin\.cursor\projects_folders\RMT\Dachser Asics\input")
PROCESSING_DIR = Path(
    r"C:\Users\avitkin\.cursor\projects_folders\RMT\Dachser Asics\processing"
)
OUTPUT_DIR = Path(r"C:\Users\avitkin\.cursor\projects_folders\RMT\Dachser Asics\output")


@dataclass(frozen=True)
class PipelinePaths:
    input_dir: Path
    processing_dir: Path
    output_dir: Path


def get_pipeline_paths() -> PipelinePaths:
    paths = PipelinePaths(
        input_dir=INPUT_DIR.resolve(),
        processing_dir=PROCESSING_DIR.resolve(),
        output_dir=OUTPUT_DIR.resolve(),
    )
    paths.processing_dir.mkdir(parents=True, exist_ok=True)
    paths.output_dir.mkdir(parents=True, exist_ok=True)
    if not paths.input_dir.is_dir():
        raise SystemExit(f"Input folder not found: {paths.input_dir}")
    return paths


def _safe_stem(path: Path) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", path.stem)


def _list_xlsx(folder: Path) -> list[Path]:
    files = sorted(folder.glob("*.xlsx"), key=lambda p: p.name.lower())
    return [p for p in files if not p.name.startswith("~$")]


def _pick_file(files: list[Path], prompt: str) -> Path:
    print(prompt)
    for i, p in enumerate(files, 1):
        print(f"  {i}. {p.name}")
    while True:
        raw = input(f"Enter number (1–{len(files)}): ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(files):
            return files[int(raw) - 1]
        print("Invalid choice.")


def _resolve_input_files(files: list[Path]) -> tuple[Path, Path]:
    if not files:
        raise SystemExit(f"No .xlsx files in {INPUT_DIR}")

    print("\nSelect input workbooks from:")
    print(f"  {INPUT_DIR}\n")
    transport = _pick_file(files, "Transport costs (main.py — Europe tariffs, toll tabs):")
    rest = [p for p in files if p.resolve() != transport.resolve()]
    if not rest:
        raise SystemExit("Need a second .xlsx for accessorial (appendix).")
    accessorial = _pick_file(rest, "Accessorial costs (main_accessorial.py — appendix):")
    print("\nSelected:")
    print(f"  Transport:   {transport.name}")
    print(f"  Accessorial: {accessorial.name}")
    return transport, accessorial


def _workbook_has_toll_tabs(path: Path) -> bool:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return any("toll" in name.lower() for name in wb.sheetnames)
    finally:
        wb.close()


def _write_json(path: Path, data: dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return path


def _export_transport_json(transport_xlsx: Path, processing_dir: Path) -> Path:
    prev_out = transport_main.OUTPUT_DIR
    transport_main.OUTPUT_DIR = processing_dir
    try:
        out = transport_main.transform_workbook(transport_xlsx)
    finally:
        transport_main.OUTPUT_DIR = prev_out
    target = processing_dir / f"{_safe_stem(transport_xlsx)}.json"
    if out.resolve() != target.resolve():
        shutil.copy2(out, target)
    print(f"  JSON: {target.name}")
    return target


def _export_accessorial_json(
    accessorial_xlsx: Path, processing_dir: Path
) -> Path:
    data = main_accessorial.parse_accessorial_workbook(accessorial_xlsx)
    target = processing_dir / f"{_safe_stem(accessorial_xlsx)}.accessorial.json"
    _write_json(target, data)
    print(f"  JSON: {target.name}")
    return target


def _export_toll_json(transport_xlsx: Path, processing_dir: Path) -> Path | None:
    if not _workbook_has_toll_tabs(transport_xlsx):
        print("  (no Toll* sheets — skipping toll JSON)")
        return None
    data = main_toll.extract_workbook_tolls(transport_xlsx)
    if not data.get("TollSheets"):
        print("  (no toll data extracted — skipping toll JSON)")
        return None
    target = processing_dir / f"{_safe_stem(transport_xlsx)}.toll.json"
    _write_json(target, data)
    print(f"  JSON: {target.name}")
    return target


def _build_transport_xlsx(transport_json: Path, work_xlsx: Path) -> Path:
    with open(transport_json, encoding="utf-8") as f:
        data = json.load(f)
    transform_json_to_xlsx(data, work_xlsx, sheet_title="Rate Card")
    print(f"  Sheet: Rate Card -> {work_xlsx.name}")
    return work_xlsx


def _build_accessorial_xlsx(
    accessorial_json: Path,
    work_xlsx: Path,
    unmatched_txt: Path,
) -> Path | None:
    included, unmatched, special = build_ra_table(accessorial_json)
    if not included:
        print("  (no accessorial rows — skipping accessorial sheet)")
        write_report(unmatched_txt, unmatched, special)
        return None
    write_xlsx(work_xlsx, included)
    wb = load_workbook(work_xlsx)
    try:
        ws = wb["RA Accessorial"]
        apply_accessorial_sheet_layout(ws)
        wb.save(work_xlsx)
    finally:
        wb.close()
    write_report(unmatched_txt, unmatched, special)
    print(f"  Sheet: RA Accessorial ({len(included)} rows) -> {work_xlsx.name}")
    return work_xlsx


def _build_toll_xlsx(
    toll_json: Path,
    work_xlsx: Path,
    postal_txt: Path,
    regions_txt: Path,
) -> Path | None:
    with open(toll_json, encoding="utf-8") as f:
        data = json.load(f)
    wb, n_tables, n_pct, postal_cat, region_cat = transform_toll_json_to_workbook(data)
    if not wb.sheetnames:
        print("  (no toll sheets — skipping toll workbook)")
        return None
    for name in wb.sheetnames:
        ws = wb[name]
        if name.lower().startswith("toll %"):
            apply_toll_pct_sheet_layout(ws)
        elif name.lower() == "toll":
            apply_toll_structured_sheet_layout(ws)
    work_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(work_xlsx)
    write_postal_zones_catalog(postal_txt, postal_cat)
    write_country_regions_catalog(regions_txt, region_cat)
    print(f"  Toll sheets ({n_tables} table(s), {n_pct} % row(s)) -> {work_xlsx.name}")
    return work_xlsx


def _merge_workbooks(target: Path, sources: list[Path]) -> None:
    out = Workbook()
    default = out.active
    if default is not None:
        out.remove(default)
    for src_path in sources:
        if not src_path.is_file():
            continue
        src = load_workbook(src_path)
        try:
            for ws in src.worksheets:
                copy_worksheet(ws, out, ws.title)
        finally:
            src.close()
    target.parent.mkdir(parents=True, exist_ok=True)
    out.save(target)


def run_pipeline() -> None:
    paths = get_pipeline_paths()
    processing_dir = paths.processing_dir
    output_dir = paths.output_dir

    files = _list_xlsx(paths.input_dir)
    print("\nPaths:")
    print(f"  input:      {paths.input_dir}")
    print(f"  processing: {processing_dir}")
    print(f"  output:     {output_dir}")

    transport_xlsx, accessorial_xlsx = _resolve_input_files(files)

    transport_stem = _safe_stem(transport_xlsx)
    work_dir = processing_dir / "_work" / transport_stem
    work_dir.mkdir(parents=True, exist_ok=True)

    final_xlsx = output_dir / f"{transport_stem}.xlsx"
    acc_unmatched = output_dir / f"{_safe_stem(accessorial_xlsx)}.unmatched.txt"
    toll_postal_out = output_dir / f"{transport_stem}.toll_postal_zones.txt"
    toll_regions_out = output_dir / f"{transport_stem}.toll_country_regions.txt"

    print("\n--- Step 1: Export JSON to processing/ ---")
    transport_json = _export_transport_json(transport_xlsx, processing_dir)
    accessorial_json = _export_accessorial_json(accessorial_xlsx, processing_dir)
    toll_json = _export_toll_json(transport_xlsx, processing_dir)

    print("\n--- Step 2: Transform JSON to Excel ---")
    merge_sources: list[Path] = []

    transport_xlsx_path = work_dir / "transport.xlsx"
    _build_transport_xlsx(transport_json, transport_xlsx_path)
    merge_sources.append(transport_xlsx_path)

    acc_xlsx_path = work_dir / "accessorial.xlsx"
    acc_built = _build_accessorial_xlsx(
        accessorial_json, acc_xlsx_path, acc_unmatched
    )
    if acc_built:
        merge_sources.append(acc_built)

    if toll_json:
        toll_xlsx_path = work_dir / "toll.xlsx"
        toll_built = _build_toll_xlsx(
            toll_json,
            toll_xlsx_path,
            toll_postal_out,
            toll_regions_out,
        )
        if toll_built:
            merge_sources.append(toll_built)

    print("\n--- Step 3: Combine sheets ---")
    _merge_workbooks(final_xlsx, merge_sources)

    print("\nDone.")
    print(f"  Combined Excel: {final_xlsx}")
    if acc_unmatched.is_file():
        print(f"  Accessorial report: {acc_unmatched}")
    if toll_postal_out.is_file():
        print(f"  Toll postal zones: {toll_postal_out}")
    if toll_regions_out.is_file():
        print(f"  Toll country regions: {toll_regions_out}")


def main() -> None:
    try:
        run_pipeline()
    except KeyboardInterrupt:
        print("\nCancelled.", file=sys.stderr)
        raise SystemExit(130) from None


if __name__ == "__main__":
    main()
