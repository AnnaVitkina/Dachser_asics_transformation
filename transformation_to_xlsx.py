"""
Transform tariff JSON (output of main.py) into a flat Excel rate-card layout.

Input: JSON with top-level MainCosts[] (source_file / standard_schema optional).
Output: .xlsx with a single table: shipment columns once, then all transport-cost
columns from every MainCosts block placed side by side (one horizontal band of
headers, one row per lane with all amounts on that row).

Does not read any template/example workbook; layout is defined only here.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = SCRIPT_DIR / "output"

SHIPMENT_HEADERS = (
    "Lane #",
    "Origin Country",
    "Origin Postal Code",
    "Destination Country",
    "Destination Postal Code",
    "Service Type",
    "Measurement",
)
BASE_COLS = len(SHIPMENT_HEADERS)  # 7; cost columns start at BASE_COLS + 1

# Same tokens as main.py (for stripping from header_context when classifying bands).
RATE_ID_TOKENS = frozenset(
    {"targospeed", "targofix", "directload", "targoflex", "zalando"}
)

KIND_KG = "kg"
KIND_PALLET = "pallet"
KIND_RETURN_30 = "return_30"
KIND_TARGOFIX = "targofix"
KIND_TARGOFIX_3 = "targofix_3"
GLOBAL_TARGOFIX3_TAB = "Targofix 3%"
JSON_COST_KINDS = frozenset({KIND_KG, KIND_PALLET})
# Per-tab synthetics (excluded from global duplicate block).
TAB_EXTRA_COST_KINDS = frozenset({KIND_RETURN_30, KIND_TARGOFIX})
EXTRA_COST_KINDS = TAB_EXTRA_COST_KINDS | {KIND_TARGOFIX_3}

# (tab_name, kind, rateid, meta) — meta may hold pad_band or rcm_index.
ColumnSpec = tuple[str, str, str | None, dict | None]


def _spec_meta(spec: ColumnSpec) -> dict:
    return spec[3] if len(spec) > 3 and spec[3] else {}


def _is_pad_column(spec: ColumnSpec) -> bool:
    return bool(_spec_meta(spec).get("pad_band"))


def _is_dup_column(spec: ColumnSpec) -> bool:
    return _spec_meta(spec).get("dup_source") is not None


def _is_global_targofix3_tab(tab: str) -> bool:
    return tab == GLOBAL_TARGOFIX3_TAB


def _parse_departure(dep: str | None) -> tuple[str | None, str | None]:
    if not dep or not str(dep).strip():
        return None, None
    s = str(dep).strip()
    m = re.match(r"^([A-Za-z]{2})-(.+)$", s)
    if m:
        return m.group(1).upper(), m.group(2).strip()
    if "-" in s:
        a, b = s.split("-", 1)
        return a.strip() or None, b.strip() or None
    return s, None


def _format_destination_postal(code: object) -> str | None:
    if code is None:
        return None
    s = str(code).strip().replace("*", "")
    if not s:
        return None
    if re.fullmatch(r"\d", s):
        return f"0{s}"
    return s


def _lane_key_has_asterisk_destination(k: tuple[str, str, str]) -> bool:
    _dep, _arr, z = k
    return z is not None and "*" in str(z)


def _service_type_for_asterisk_lane(
    k: tuple[str, str, str],
    segments: list[dict],
) -> str | None:
    """e.g. TARGOFLEX KG for lanes whose JSON destination_zip_code contains '*'."""
    if not _lane_key_has_asterisk_destination(k):
        return None
    for s in segments:
        for dr in s["rows"]:
            if _lane_key(dr) != k:
                continue
            z = dr.get("destination_zip_code")
            if z is None or "*" not in str(z):
                continue
            rids = _segment_rate_ids(s)
            token = (rids[0] if rids else None) or (
                str(s.get("rateid") or "").strip().lower() or None
            )
            if not token:
                return None
            return f"{token.upper()} KG"
    return None


def _strip_pallet_amounts_for_asterisk_lanes(
    amounts: dict[tuple[str, str, str], list[object | None]],
    column_specs: list[ColumnSpec],
) -> None:
    """Lanes with '*' in destination zip: no per-pallet rates (incl. pad cols and dup of pallet)."""
    for k, row_amt in amounts.items():
        if not _lane_key_has_asterisk_destination(k):
            continue
        for i, spec in enumerate(column_specs):
            kind = spec[1]
            meta = _spec_meta(spec)
            if kind == KIND_PALLET:
                row_amt[i] = None
                continue
            if meta.get("pad_band"):
                row_amt[i] = None
                continue
            src = meta.get("dup_source")
            if src is not None:
                src_spec = column_specs[src]
                if src_spec[1] == KIND_PALLET or _spec_meta(src_spec).get("pad_band"):
                    row_amt[i] = None


def _is_noise_tariff_row(dr: dict) -> bool:
    """Zalando footer address lines and sheet title rows — not tariff lanes."""
    dep = str(dr.get("departure") or "").strip()
    if not dep:
        return False
    low = dep.lower()
    arr = dr.get("arrival_country")
    z = dr.get("destination_zip_code")
    arr_empty = arr is None or not str(arr).strip()
    z_empty = z is None or not str(z).strip()
    if low.startswith("zalando ") and arr_empty and z_empty:
        return True
    if "only for this consignee" in low:
        return True
    if "tariff" in low and "consignee" in low:
        return True
    return False


def _destination_zip_two_digit_prefix(z: object) -> str | None:
    if z is None:
        return None
    digits = re.sub(r"\D", "", str(z).replace("*", ""))
    if not digits:
        return None
    if len(digits) == 1:
        return f"0{digits}"
    return digits[:2]


def _arrival_is_poland(arr: str) -> bool:
    a = (arr or "").strip().lower()
    return a in ("poland", "pl", "p")


def _arrival_is_germany(arr: str) -> bool:
    a = (arr or "").strip().lower()
    return a in ("germany", "d", "de", "deutschland")


ZALANDO_CONSIGNEE_PREFIXES_PL = frozenset({"72", "95"})
ZALANDO_CONSIGNEE_PREFIXES_DE = frozenset({"06", "14", "41", "77", "99"})


def _shard_matches_zalando_consignee(shard: dict) -> bool:
    arr = str(shard.get("arrival_country") or "").strip()
    pref = _destination_zip_two_digit_prefix(shard.get("destination_zip_code"))
    if not pref:
        return False
    if _arrival_is_poland(arr):
        return pref in ZALANDO_CONSIGNEE_PREFIXES_PL
    if _arrival_is_germany(arr):
        return pref in ZALANDO_CONSIGNEE_PREFIXES_DE
    return False


def _zalando_main_spec_range(
    segments: list[dict], spec_offsets: list[int], main_spec_count: int
) -> tuple[int, int] | None:
    for si, s in enumerate(segments):
        if "zalando" not in str(s.get("tab") or "").lower():
            continue
        start = spec_offsets[si]
        end = (
            spec_offsets[si + 1]
            if si + 1 < len(spec_offsets)
            else main_spec_count
        )
        return (start, end)
    return None


def _row_has_amount_in_range(
    row_amt: list[object | None], lo: int, hi: int
) -> bool:
    for i in range(lo, hi):
        v = row_amt[i]
        if v is None:
            continue
        if isinstance(v, str) and not str(v).strip():
            continue
        return True
    return False


def _service_type_cell_value(
    k: tuple[str, str, str],
    segments: list[dict],
    row_amt: list[object | None],
    main_spec_count: int,
    spec_offsets: list[int],
    shard: dict,
    *,
    include_zalando: bool = True,
) -> str | None:
    st = _service_type_for_asterisk_lane(k, segments)
    if st:
        return st
    if not include_zalando:
        return None
    zr = _zalando_main_spec_range(segments, spec_offsets, main_spec_count)
    if zr is None:
        return None
    lo, hi = zr
    if not _row_has_amount_in_range(row_amt, lo, hi):
        return None
    if not _shard_matches_zalando_consignee(shard):
        return None
    return "Zalando"


def _ctx_tuple(entry: dict) -> tuple[str, ...]:
    ctx = entry.get("header_context") or []
    return tuple(str(x).strip() for x in ctx if x is not None and str(x).strip())


def _rateid_contains(rateid: str | None, token: str) -> bool:
    return token.lower() in (rateid or "").lower()


def _segment_rate_ids(segment: dict) -> list[str]:
    """All rate ids for a tab (rateids[] from JSON, else legacy single rateid)."""
    out: list[str] = []
    seen: set[str] = set()
    raw = segment.get("rateids")
    if isinstance(raw, list):
        for x in raw:
            t = str(x).strip().lower()
            if t and t not in seen:
                seen.add(t)
                out.append(t)
    rid = segment.get("rateid")
    if isinstance(rid, str) and rid.strip():
        t = rid.strip().lower()
        if t not in seen:
            seen.add(t)
            out.insert(0, t)
    if not out:
        for entry in segment.get("rcm") or []:
            for token in _rate_ids_from_header_context(entry.get("header_context")):
                if token not in seen:
                    seen.add(token)
                    out.append(token)
    return out


def _rateids_match_token(rateids: list[str] | None, token: str) -> bool:
    return any(_rateid_contains(r, token) for r in (rateids or []))


def _find_rate_id_in_text(text: str) -> str | None:
    low = str(text).lower()
    for token in RATE_ID_TOKENS:
        if re.search(
            rf"(?i)(?<![a-z0-9]){re.escape(token)}(?![a-z0-9])",
            low,
        ):
            return token
    return None


def _rate_ids_from_header_context(ctx: list | None) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for x in ctx or []:
        token = _find_rate_id_in_text(str(x))
        if token and token not in seen:
            seen.add(token)
            found.append(token)
    return found


def _rate_ids_for_spec_block(
    spec_start: int,
    spec_end: int,
    column_specs: list[ColumnSpec],
    segments: list[dict],
    spec_offsets: list[int],
    spec_idx: int,
) -> list[str]:
    """Rate ids for a (tab, kind) block: column header_context first, then tab rateids."""
    found: list[str] = []
    seen: set[str] = set()
    for i in range(spec_start, spec_end):
        if column_specs[i][1] not in JSON_COST_KINDS:
            continue
        entry = _rcm_entry_for_spec(i, column_specs, segments, spec_offsets)
        if not entry:
            continue
        for rid in _rate_ids_from_header_context(entry.get("header_context")):
            if rid not in seen:
                seen.add(rid)
                found.append(rid)
    if found:
        return found
    si, _, _ = _segment_spec_range(spec_idx, segments, spec_offsets, column_specs)
    if si is not None:
        return _segment_rate_ids(segments[si])
    return []


def _rateid_is_return_30(rateids: list[str] | None) -> bool:
    """Return 30% column when rateid is targoflex."""
    return _rateids_match_token(rateids, "targoflex")


def _extra_kinds_for_segment(rateids: list[str] | None) -> list[str]:
    """Synthetic columns appended once at end of tab (Return 30% is per cost block)."""
    extras: list[str] = []
    if _rateids_match_token(rateids, "targofix"):
        extras.append(KIND_TARGOFIX)
    return extras


def _tabs_with_mixed_cost_kinds(
    column_specs: list[tuple[str, str, str | None]],
) -> dict[str, bool]:
    by_tab: dict[str, set[str]] = {}
    for spec in column_specs:
        tab, kind = spec[0], spec[1]
        if kind not in JSON_COST_KINDS:
            continue
        by_tab.setdefault(tab, set()).add(kind)
    return {tab: len(kinds) > 1 for tab, kinds in by_tab.items()}


def _tab_has_extra_columns(
    column_specs: list[tuple[str, str, str | None]],
    tab_start: int,
    tab_end: int,
) -> bool:
    return any(column_specs[i][1] in TAB_EXTRA_COST_KINDS for i in range(tab_start, tab_end))


def _any_tab_needs_grouped_row(
    column_specs: list[tuple[str, str, str | None]],
) -> bool:
    tabs_mixed = _tabs_with_mixed_cost_kinds(column_specs)
    i = 0
    n = len(column_specs)
    while i < n:
        tab = column_specs[i][0]
        j = i + 1
        while j < n and column_specs[j][0] == tab:
            j += 1
        if _is_global_targofix3_tab(tab):
            return True
        if tabs_mixed.get(tab, False) or _tab_has_extra_columns(column_specs, i, j):
            return True
        i = j
    return False


def _strip_rate_ids_from_context(ctx: list[str]) -> list[str]:
    out: list[str] = []
    for x in ctx:
        t = str(x).strip()
        if t.lower() in RATE_ID_TOKENS:
            continue
        out.append(t)
    return out


def _parse_num_token(text: str) -> float | None:
    try:
        return float(str(text).replace(",", ".").replace(" ", ""))
    except ValueError:
        return None


def _is_weight_lower_bound_token(text: str) -> bool:
    """e.g. 29.1, 59.1 — lower edge of a kg band, not a pallet count."""
    return bool(re.match(r"^\d+\.1$", str(text).strip()))


def _is_flat_rate_pallet_context(joined: str) -> bool:
    return "flat rate" in joined or "flate rate" in joined


def _excel_header_is_numeric_cap(eh: str) -> float | None:
    s = eh.strip()
    if re.fullmatch(r"\d+(?:\.\d+)?", s):
        return _parse_num_token(s)
    return None


def _classify_cost_column(entry: dict) -> str:
    """Return 'kg' or 'pallet'."""
    eh = str(entry.get("excel_header") or "")
    eh_l = eh.lower()
    ctx = [str(x) for x in (entry.get("header_context") or []) if x is not None]
    joined = " ".join(c.lower() for c in ctx)
    stripped = _strip_rate_ids_from_context(ctx)

    if "pallet" in eh_l or "pallet" in joined:
        return "pallet"
    if re.search(r"\d+\s*to\s*\d+", eh, re.IGNORECASE) and "pallet" in eh_l:
        return "pallet"
    if _is_flat_rate_pallet_context(joined):
        return "pallet"

    eh_num = _excel_header_is_numeric_cap(eh)
    has_per_shipment = "per shipment" in eh_l or any(
        "per shipment" in str(x).lower() for x in ctx
    )
    has_weight_lower = any(_is_weight_lower_bound_token(t) for t in stripped)
    has_small_start = any(
        (n := _parse_num_token(t)) is not None and n < 30 for t in stripped
    )

    # Weight: numeric cap in excel_header (29 / 59 / 100) with Per shipment or band context.
    if eh_num is not None and (has_per_shipment or has_weight_lower or has_small_start):
        if (
            not has_per_shipment
            and has_weight_lower
            and len(stripped) == 1
            and eh_num <= 33
            and eh_num < 29
        ):
            return "pallet"
        if has_per_shipment or has_weight_lower or eh_num >= 29:
            return "kg"

    if eh_l.startswith("per shipment") and not _is_flat_rate_pallet_context(joined):
        return "kg"

    nums: list[float] = []
    for t in stripped:
        n = _parse_num_token(t)
        if n is not None:
            nums.append(n)
    if len(stripped) >= 2 and len(nums) >= 2:
        return "kg"
    if has_weight_lower and eh_num is not None and eh_num >= 29:
        return "kg"

    if eh_num is not None and eh_num == int(eh_num) and int(eh_num) <= 33:
        return "pallet"
    return "pallet"


def _parse_two_ints_from_to_pallets(text: str) -> tuple[int | None, int | None]:
    m = re.search(
        r"(\d+(?:\.\d+)?)\s*to\s*(\d+(?:\.\d+)?)\s*pallets?",
        text,
        re.IGNORECASE,
    )
    if not m:
        return None, None
    try:
        return int(float(m.group(1))), int(float(m.group(2)))
    except ValueError:
        return None, None


def _band_label(entry: dict, kind: str) -> str:
    ctx_raw = [str(x) for x in (entry.get("header_context") or []) if x is not None]
    ctx = _strip_rate_ids_from_context(ctx_raw)
    eh = str(entry.get("excel_header") or "")

    if kind == "pallet":
        a, b = _parse_two_ints_from_to_pallets(eh)
        if b is not None:
            return f"<={b}"
        eh_num = _excel_header_is_numeric_cap(eh)
        if eh_num is not None and eh_num == int(eh_num):
            return f"<={int(eh_num)}"
        for t in ctx:
            if _is_weight_lower_bound_token(t):
                continue
            m = re.match(r"^(\d+(?:\.\d+)?)\s*$", t.strip())
            if m:
                try:
                    v = float(m.group(1))
                    if v == int(v):
                        return f"<={int(v)}"
                    return f"<={v}"
                except ValueError:
                    pass
        return "<=0"

    eh_num = _excel_header_is_numeric_cap(eh)
    if eh_num is not None:
        if eh_num == int(eh_num):
            return f"<={int(eh_num)}"
        return f"<={eh_num}"

    nums: list[float] = []
    for t in ctx:
        n = _parse_num_token(t)
        if n is not None:
            nums.append(n)
    if nums:
        last = nums[-1]
        if last == int(last):
            return f"<={int(last)}"
        return f"<={last}"
    return "<=0"


def _band_upper_bound(entry: dict, kind: str) -> int | None:
    """Numeric upper bound from a <=N band label."""
    m = re.match(r"<=\s*(\d+(?:\.\d+)?)", _band_label(entry, kind))
    if not m:
        return None
    v = float(m.group(1))
    return int(v) if v == int(v) else int(v)


def _cost_name_cell(tab_name: str, kind: str) -> str:
    """Tab + per kg or per Pallets only (no numeric band / rate-id text)."""
    tab = str(tab_name or "").strip()
    if kind == "pallet":
        return f"Transport cost ({tab} / per Pallets)"
    return f"Transport cost ({tab} / per kg)"


def _grouped_cost_tab_label(tab_name: str) -> str:
    tab = str(tab_name or "").strip()
    return f"Grouped cost: Transport cost ({tab})"


def _tab_has_per_block_return_30(
    column_specs: list[ColumnSpec],
    tab_start: int,
    tab_end: int,
) -> bool:
    return any(
        column_specs[i][1] == KIND_RETURN_30
        and _spec_meta(column_specs[i]).get("return_30_for_kind")
        for i in range(tab_start, tab_end)
    )


def _grouped_cost_split_blocks(
    column_specs: list[ColumnSpec],
    tab_start: int,
    tab_end: int,
) -> list[tuple[str, int, int]]:
    """Per (kg|pallet) block: JSON columns + matching Return 30% column."""
    blocks: list[tuple[str, int, int]] = []
    i = tab_start
    while i < tab_end:
        kind = column_specs[i][1]
        if kind not in JSON_COST_KINDS:
            i += 1
            continue
        json_kind = kind
        j = i + 1
        while j < tab_end and column_specs[j][1] == json_kind:
            j += 1
        block_end = j
        if block_end < tab_end and column_specs[block_end][1] == KIND_RETURN_30:
            meta = _spec_meta(column_specs[block_end])
            if meta.get("return_30_for_kind") == json_kind:
                block_end += 1
        blocks.append((json_kind, i, block_end))
        i = block_end
    return blocks


def _layout_amount_pos_for_spec(
    spec_idx: int,
    tab_start: int,
    tab_layout_start: int,
    column_specs: list[ColumnSpec],
    tab_end: int,
) -> int:
    pos = tab_layout_start
    if _tab_has_json_columns(column_specs, tab_start, tab_end):
        pos += 1
    return pos + (spec_idx - tab_start)


def _grouped_cost_label_for_kind_block(tab: str, json_kind: str) -> str:
    return f"Grouped cost: {_cost_name_cell(tab, json_kind)}"


def _grouped_cost_label_for_tab(
    tab: str,
    column_specs: list[tuple[str, str, str | None]],
    tab_start: int,
    tab_end: int,
    tabs_mixed: dict[str, bool],
) -> str:
    if _is_global_targofix3_tab(tab):
        return "Grouped cost: Transport cost (Targofix 3%)"
    if tabs_mixed.get(tab, False) and not _tab_has_per_block_return_30(
        column_specs, tab_start, tab_end
    ):
        return _grouped_cost_tab_label(tab)
    last_json_kind: str | None = None
    for i in range(tab_start, tab_end):
        k = column_specs[i][1]
        if k in JSON_COST_KINDS:
            last_json_kind = k
    if last_json_kind:
        return f"Grouped cost: {_cost_name_cell(tab, last_json_kind)}"
    return _grouped_cost_tab_label(tab)


def _cost_name_for_kind(tab_name: str, kind: str) -> str:
    if kind == KIND_RETURN_30:
        return "Transport cost (Return, 30%)"
    if kind == KIND_TARGOFIX:
        return "Transport cost (Targofix)"
    if kind == KIND_TARGOFIX_3:
        return "Transport cost (Targofix 3%)"
    return _cost_name_cell(tab_name, kind)


def _display_tab_kind_rateid(
    spec_idx: int, column_specs: list[ColumnSpec]
) -> tuple[str, str, str | None]:
    spec = column_specs[spec_idx]
    src = _spec_meta(spec).get("dup_source")
    if src is not None:
        spec = column_specs[src]
    return spec[0], spec[1], spec[2]


def _cost_name_for_spec_idx(spec_idx: int, column_specs: list[ColumnSpec]) -> str:
    spec = column_specs[spec_idx]
    kind = spec[1]
    if kind == KIND_RETURN_30:
        return "Transport cost (Return, 30%)"
    if kind == KIND_TARGOFIX:
        return "Transport cost (Targofix)"
    if kind == KIND_TARGOFIX_3:
        return "Transport cost (Targofix 3%)"
    tab, kind, _rid = _display_tab_kind_rateid(spec_idx, column_specs)
    return _cost_name_cell(tab, kind)


def _applies_over_global_targofix3(column_specs: list[ColumnSpec]) -> str:
    seen: list[str] = []
    for spec in column_specs:
        src = _spec_meta(spec).get("dup_source")
        if src is None:
            continue
        s = column_specs[src]
        name = _cost_name_cell(s[0], s[1])
        if name not in seen:
            seen.append(name)
    return "Applies over " + "; ".join(seen)


def _json_kinds_for_tab(
    column_specs: list[tuple[str, str, str | None]],
    tab_start: int,
    tab_end: int,
) -> list[str]:
    """JSON cost kinds for a tab (pallet before kg when both present)."""
    kinds: list[str] = []
    for i in range(tab_start, tab_end):
        k = column_specs[i][1]
        if k in JSON_COST_KINDS and k not in kinds:
            kinds.append(k)
    ordered: list[str] = []
    if KIND_PALLET in kinds:
        ordered.append(KIND_PALLET)
    if KIND_KG in kinds:
        ordered.append(KIND_KG)
    for k in kinds:
        if k not in ordered:
            ordered.append(k)
    return ordered


def _applies_over_label(
    tab: str,
    column_specs: list[tuple[str, str, str | None]],
    tab_start: int,
    tab_end: int,
) -> str:
    names = [
        _cost_name_cell(tab, k) for k in _json_kinds_for_tab(column_specs, tab_start, tab_end)
    ]
    return "Applies over " + "; ".join(names)


def _tab_spec_range(
    column_specs: list[tuple[str, str, str | None]], tab: str
) -> tuple[int, int]:
    start = next(i for i, spec in enumerate(column_specs) if spec[0] == tab)
    end = start + 1
    while end < len(column_specs) and column_specs[end][0] == tab:
        end += 1
    return start, end


def _applies_if_for_spec_idx(
    spec_idx: int,
    column_specs: list[ColumnSpec],
    segments: list[dict],
    spec_offsets: list[int],
) -> str:
    spec = column_specs[spec_idx]
    kind = spec[1]
    if kind == KIND_TARGOFIX_3:
        return _applies_over_global_targofix3(column_specs)
    if kind == KIND_RETURN_30:
        meta = _spec_meta(spec)
        for_kind = meta.get("return_30_for_kind")
        if for_kind:
            return f"Applies over {_cost_name_cell(spec[0], for_kind)}"
        tab_start, tab_end = _tab_spec_range(column_specs, spec[0])
        return _applies_over_label(spec[0], column_specs, tab_start, tab_end)
    if kind == KIND_TARGOFIX:
        tab = spec[0]
        tab_start, tab_end = _tab_spec_range(column_specs, tab)
        return _applies_over_label(tab, column_specs, tab_start, tab_end)
    block_start = spec_idx
    while block_start > 0 and column_specs[block_start - 1][1] == kind:
        block_start -= 1
    block_end = spec_idx + 1
    while block_end < len(column_specs) and column_specs[block_end][1] == kind:
        block_end += 1
    rateids = _rate_ids_for_spec_block(
        block_start, block_end, column_specs, segments, spec_offsets, spec_idx
    )
    return _applies_if_lane(rateids, kind)


def _tab_has_json_columns(
    column_specs: list[tuple[str, str, str | None]],
    tab_start: int,
    tab_end: int,
) -> bool:
    return any(column_specs[i][1] in JSON_COST_KINDS for i in range(tab_start, tab_end))


def _tab_layout_width(
    column_specs: list[tuple[str, str, str | None]],
    tab_start: int,
    tab_end: int,
) -> int:
    """One currency column at tab start (when JSON costs exist) + all amount columns."""
    n_amounts = tab_end - tab_start
    if _tab_has_json_columns(column_specs, tab_start, tab_end):
        return 1 + n_amounts
    return n_amounts


def _value_for_spec_column(
    column_specs: list[tuple[str, str, str | None]],
    spec_index: int,
    measurement: str,
    row_vals: list[object | None],
) -> object | None:
    spec = column_specs[spec_index]
    kind = spec[1]
    if kind == KIND_RETURN_30:
        return 30 if measurement == "Return" else None
    if kind in (KIND_TARGOFIX, KIND_TARGOFIX_3):
        return 3
    meta = _spec_meta(column_specs[spec_index])
    src = meta.get("dup_source")
    if src is not None:
        return row_vals[src]
    return row_vals[spec_index]


def _is_request_value(val: object) -> bool:
    if val is None or not isinstance(val, str):
        return False
    low = val.strip().lower()
    return low in ("on request", "upon request") or "request" in low


def _has_meaningful_cost_value(val: object) -> bool:
    """Non-empty cell that is not a request placeholder."""
    if val is None:
        return False
    if _is_request_value(val):
        return False
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return False
        try:
            float(s.replace(",", ".").replace(" ", ""))
            return True
        except ValueError:
            return True
    return True


def _cost_segment_ranges(
    column_specs: list[tuple[str, str, str | None]],
) -> list[tuple[int, int]]:
    """(start, end) spec indices per (tab, kind) block; extras are one column each."""
    ranges: list[tuple[int, int]] = []
    i = 0
    n = len(column_specs)
    while i < n:
        tab = column_specs[i][0]
        kind = column_specs[i][1]
        j = i + 1
        while j < n and column_specs[j][0] == tab and column_specs[j][1] == kind:
            j += 1
        ranges.append((i, j))
        i = j
    return ranges


def _normalize_request_values_in_row(
    row_vals: list[object | None],
    column_specs: list[tuple[str, str, str | None]],
) -> None:
    """
    Per (tab, kind) segment: request cells -> 0 if any other value exists, else empty.
    """
    for start, end in _cost_segment_ranges(column_specs):
        if column_specs[start][1] in EXTRA_COST_KINDS or _is_global_targofix3_tab(
            column_specs[start][0]
        ):
            continue
        request_indices = [
            i for i in range(start, end) if _is_request_value(row_vals[i])
        ]
        if not request_indices:
            continue
        has_other = any(
            _has_meaningful_cost_value(row_vals[i])
            for i in range(start, end)
            if not _is_pad_column(column_specs[i])
        )
        for i in request_indices:
            row_vals[i] = 0 if has_other else None


def _append_global_targofix3_section(
    specs: list[ColumnSpec],
) -> list[ColumnSpec]:
    """Duplicate all JSON/pad columns; append Transport cost (Targofix 3%)."""
    out = list(specs)
    base_len = len(specs)
    for i in range(base_len):
        if specs[i][1] in TAB_EXTRA_COST_KINDS:
            continue
        src = specs[i]
        out.append((GLOBAL_TARGOFIX3_TAB, src[1], src[2], {"dup_source": i}))
    out.append((GLOBAL_TARGOFIX3_TAB, KIND_TARGOFIX_3, None, None))
    return out


def _build_main_column_specs(segments: list[dict]) -> list[ColumnSpec]:
    """One spec per global cost column; may insert pallet pad columns (<=N-1, value 0)."""
    specs: list[ColumnSpec] = []
    for s in segments:
        tab = s["tab"]
        rateids = _segment_rate_ids(s)
        rateid = rateids[0] if rateids else s.get("rateid")
        rcm = s["rcm"]
        i = 0
        while i < len(rcm):
            kind = _classify_cost_column(rcm[i])
            j = i + 1
            while j < len(rcm) and _classify_cost_column(rcm[j]) == kind:
                j += 1
            if kind == KIND_PALLET:
                first_ub = _band_upper_bound(rcm[i], kind)
                if first_ub is not None and first_ub > 1:
                    specs.append(
                        (tab, kind, rateid, {"pad_band": f"<={first_ub - 1}"})
                    )
            for k in range(i, j):
                specs.append((tab, kind, rateid, {"rcm_index": k}))
            if _rateid_is_return_30(rateids):
                specs.append(
                    (tab, KIND_RETURN_30, rateid, {"return_30_for_kind": kind})
                )
            i = j
        for extra_kind in _extra_kinds_for_segment(rateids):
            specs.append((tab, extra_kind, rateid, None))
    return specs


def _build_column_specs(segments: list[dict]) -> list[ColumnSpec]:
    return _append_global_targofix3_section(_build_main_column_specs(segments))


def _fill_global_section_columns(
    row_vals: list[object | None],
    column_specs: list[ColumnSpec],
    main_spec_count: int,
) -> None:
    for spec_idx in range(main_spec_count, len(column_specs)):
        spec = column_specs[spec_idx]
        meta = _spec_meta(spec)
        src = meta.get("dup_source")
        if src is not None:
            row_vals[spec_idx] = row_vals[src]
        elif spec[1] == KIND_TARGOFIX_3:
            row_vals[spec_idx] = 3


def _spec_offsets_for_segments(
    segments: list[dict], column_specs: list[ColumnSpec]
) -> list[int]:
    """Start index in column_specs for each segment (one tab per segment)."""
    offsets: list[int] = []
    pos = 0
    for s in segments:
        offsets.append(pos)
        tab = s["tab"]
        n = 0
        while pos + n < len(column_specs) and column_specs[pos + n][0] == tab:
            n += 1
        pos += n
    return offsets


def _segment_spec_range(
    spec_idx: int,
    segments: list[dict],
    spec_offsets: list[int],
    column_specs: list[ColumnSpec],
) -> tuple[int, int, int] | tuple[None, None, None]:
    """Return (segment_index, range_start, range_end) for a global spec index."""
    for si, s in enumerate(segments):
        start = spec_offsets[si]
        end = start
        while end < len(column_specs) and column_specs[end][0] == s["tab"]:
            end += 1
        if start <= spec_idx < end:
            return si, start, end
    return None, None, None


def _rcm_entry_for_spec(
    spec_idx: int,
    column_specs: list[ColumnSpec],
    segments: list[dict],
    spec_offsets: list[int],
) -> dict | None:
    meta = _spec_meta(column_specs[spec_idx])
    src = meta.get("dup_source")
    if src is not None:
        return _rcm_entry_for_spec(src, column_specs, segments, spec_offsets)
    if meta.get("pad_band"):
        return None
    rcm_index = meta.get("rcm_index")
    if rcm_index is None:
        return None
    si, _, _ = _segment_spec_range(
        spec_idx, segments, spec_offsets, column_specs
    )
    if si is None:
        return None
    return segments[si]["rcm"][rcm_index]


def _fill_amounts_from_segments(
    amounts: dict[tuple[str, str, str], list[object | None]],
    segments: list[dict],
    column_specs: list[ColumnSpec],
    spec_offsets: list[int],
) -> None:
    for si, s in enumerate(segments):
        start = spec_offsets[si]
        end = start
        while end < len(column_specs) and column_specs[end][0] == s["tab"]:
            end += 1
        rcm = s["rcm"]
        for dr in s["rows"]:
            k = _lane_key(dr)
            if k not in amounts:
                continue
            row_amt = amounts[k]
            for spec_idx in range(start, end):
                meta = _spec_meta(column_specs[spec_idx])
                if meta.get("pad_band"):
                    row_amt[spec_idx] = 0
                    continue
                rcm_index = meta.get("rcm_index")
                if rcm_index is None:
                    continue
                entry = rcm[rcm_index]
                key = entry.get("amount_key")
                row_amt[spec_idx] = dr.get(key) if key else None


def _build_display_layout(
    column_specs: list[tuple[str, str, str | None]],
) -> list[dict]:
    """
    One currency spacer at the start of each tab (when it has JSON costs), then all amounts.
    Items: {"type": "currency"} or {"type": "amount", "amount_index": int}.
    """
    layout: list[dict] = []
    i = 0
    n = len(column_specs)
    while i < n:
        tab = column_specs[i][0]
        j = i + 1
        while j < n and column_specs[j][0] == tab:
            j += 1
        if _tab_has_json_columns(column_specs, i, j):
            layout.append({"type": "currency", "tab": tab})
        for k in range(i, j):
            layout.append({"type": "amount", "amount_index": k})
        i = j
    return layout


def _layout_excel_col(layout_index: int) -> int:
    return BASE_COLS + 1 + layout_index


def _cols_for_amount_index_range(
    layout: list[dict], amount_lo: int, amount_hi_exclusive: int
) -> list[int]:
    """Excel columns for currency spacers + amounts belonging to amount index range."""
    cols: list[int] = []
    li = 0
    while li < len(layout):
        if layout[li]["type"] == "currency":
            has_amount = False
            j = li + 1
            while j < len(layout) and layout[j]["type"] == "amount":
                idx = layout[j]["amount_index"]
                if amount_lo <= idx < amount_hi_exclusive:
                    has_amount = True
                    break
                j += 1
            if has_amount:
                cols.append(_layout_excel_col(li))
            li += 1
            continue
        if layout[li]["type"] == "amount":
            idx = layout[li]["amount_index"]
            if amount_lo <= idx < amount_hi_exclusive:
                cols.append(_layout_excel_col(li))
        li += 1
    return cols


def _write_merged_row_by_tab_kind(
    ws,
    row: int,
    column_specs: list[tuple[str, str, str | None]],
    layout: list[dict],
    value_fn,
) -> None:
    """One merged cell per (tab, kind) block; one currency column at each tab start only."""
    i = 0
    n = len(column_specs)
    layout_pos = 0
    while i < n:
        tab = column_specs[i][0]
        tab_end = i + 1
        while tab_end < n and column_specs[tab_end][0] == tab:
            tab_end += 1
        if _tab_has_json_columns(column_specs, i, tab_end):
            layout_pos += 1  # single currency column for this tab
        k = i
        while k < tab_end:
            kind = column_specs[k][1]
            rateid = column_specs[k][2]
            kk = k + 1
            while kk < tab_end and column_specs[kk][1] == kind:
                kk += 1
            c_start = _layout_excel_col(layout_pos)
            c_end = _layout_excel_col(layout_pos + (kk - k) - 1)
            cell = ws.cell(row=row, column=c_start, value=value_fn(k))
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            if c_end > c_start:
                ws.merge_cells(
                    start_row=row,
                    start_column=c_start,
                    end_row=row,
                    end_column=c_end,
                )
            layout_pos += kk - k
            k = kk
        i = tab_end


def _write_merged_cost_name_row(
    ws,
    row: int,
    column_specs: list[tuple[str, str, str | None]],
    layout: list[dict],
) -> None:
    _write_merged_row_by_tab_kind(
        ws,
        row,
        column_specs,
        layout,
        lambda block_start: _cost_name_for_spec_idx(block_start, column_specs),
    )


def _applies_if_lane(rateids: list[str] | str | None, kind: str) -> str:
    if isinstance(rateids, str):
        ids = [rateids.strip().lower()] if rateids.strip() else []
    else:
        ids = [r for r in (rateids or []) if r]
    suf = "PAL" if kind == "pallet" else "KG"
    if not ids:
        return f"Applies if: Service equals {suf}"
    if len(ids) == 1:
        return f"Applies if: Service equals {ids[0]} {suf}"
    services = " or ".join(ids)
    return f"Applies if: Service equals {services} {suf}"


def _write_merged_applies_if_row(
    ws,
    row: int,
    column_specs: list[tuple[str, str, str | None]],
    layout: list[dict],
    segments: list[dict],
    spec_offsets: list[int],
) -> None:
    _write_merged_row_by_tab_kind(
        ws,
        row,
        column_specs,
        layout,
        lambda block_start: _applies_if_for_spec_idx(
            block_start, column_specs, segments, spec_offsets
        ),
    )


def _write_flat_currency_row(
    ws,
    row: int,
    layout: list[dict],
    column_specs: list[tuple[str, str, str | None]],
) -> None:
    for li, item in enumerate(layout):
        col = _layout_excel_col(li)
        if item["type"] == "currency":
            ws.cell(row=row, column=col, value="Currency")
        else:
            kind = column_specs[item["amount_index"]][1]
            if kind in EXTRA_COST_KINDS:
                cell = ws.cell(row=row, column=col, value="%\nOver costs")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            else:
                ws.cell(row=row, column=col, value="Flat")


def _advance_layout_pos_for_tab_block(
    column_specs: list[tuple[str, str, str | None]],
    tab_start: int,
    tab_end: int,
    layout_pos: int,
) -> int:
    return layout_pos + _tab_layout_width(column_specs, tab_start, tab_end)


def _first_amount_layout_pos_for_tab(
    column_specs: list[tuple[str, str, str | None]],
    tab_start: int,
    tab_end: int,
    tab_layout_start: int,
) -> int:
    if _tab_has_json_columns(column_specs, tab_start, tab_end):
        return tab_layout_start + 1
    return tab_layout_start


def _write_grouped_cost_merge(
    ws,
    row: int,
    c_start: int,
    c_end: int,
    label: str,
) -> None:
    cell = ws.cell(row=row, column=c_start, value=label)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if c_end > c_start:
        ws.merge_cells(
            start_row=row,
            start_column=c_start,
            end_row=row,
            end_column=c_end,
        )


def _write_grouped_cost_row(
    ws,
    row: int,
    column_specs: list[tuple[str, str, str | None]],
    layout: list[dict],
) -> None:
    """Grouped label per tab when mixed kg/pallet and/or synthetic columns were added."""
    tabs_mixed = _tabs_with_mixed_cost_kinds(column_specs)
    i = 0
    n = len(column_specs)
    layout_pos = 0
    while i < n:
        tab = column_specs[i][0]
        j = i + 1
        while j < n and column_specs[j][0] == tab:
            j += 1
        tab_layout_start = layout_pos
        layout_pos = _advance_layout_pos_for_tab_block(column_specs, i, j, layout_pos)
        needs = (
            _is_global_targofix3_tab(tab)
            or tabs_mixed.get(tab, False)
            or _tab_has_extra_columns(column_specs, i, j)
        )
        if not needs:
            i = j
            continue

        split = (
            tabs_mixed.get(tab, False)
            and _tab_has_per_block_return_30(column_specs, i, j)
            and not _is_global_targofix3_tab(tab)
        )
        if split:
            blocks = _grouped_cost_split_blocks(column_specs, i, j)
            for json_kind, b_start, b_end in blocks:
                lp_start = _layout_amount_pos_for_spec(
                    b_start, i, tab_layout_start, column_specs, j
                )
                lp_end = _layout_amount_pos_for_spec(
                    b_end - 1, i, tab_layout_start, column_specs, j
                )
                _write_grouped_cost_merge(
                    ws,
                    row,
                    _layout_excel_col(lp_start),
                    _layout_excel_col(lp_end),
                    _grouped_cost_label_for_kind_block(tab, json_kind),
                )
            last_block_end = blocks[-1][2] if blocks else i
            if last_block_end < j and _tab_has_extra_columns(
                column_specs, last_block_end, j
            ):
                lp_start = _layout_amount_pos_for_spec(
                    last_block_end, i, tab_layout_start, column_specs, j
                )
                _write_grouped_cost_merge(
                    ws,
                    row,
                    _layout_excel_col(lp_start),
                    _layout_excel_col(layout_pos - 1),
                    _grouped_cost_label_for_tab(
                        tab, column_specs, i, j, tabs_mixed
                    ),
                )
        else:
            first_amt = _first_amount_layout_pos_for_tab(
                column_specs, i, j, tab_layout_start
            )
            label = _grouped_cost_label_for_tab(
                tab, column_specs, i, j, tabs_mixed
            )
            _write_grouped_cost_merge(
                ws,
                row,
                _layout_excel_col(first_amt),
                _layout_excel_col(layout_pos - 1),
                label,
            )
        i = j


def _rate_by_lane(kind: str) -> str:
    return "Pallet/Pallet" if kind == "pallet" else "Weight/kg"


def _lane_key(dr: dict) -> tuple[str, str, str]:
    """Align rows across tabs: same departure, arrival, destination zip."""
    dep = dr.get("departure")
    dep_s = "" if dep is None else str(dep).strip()
    arr = dr.get("arrival_country")
    arr_s = "" if arr is None else str(arr).strip()
    z = dr.get("destination_zip_code")
    z_s = "" if z is None else str(z).strip()
    return (dep_s, arr_s, z_s)


def _shipment_shard(dr: dict) -> dict[str, object | None]:
    dep = dr.get("departure")
    oc, op = _parse_departure(dep if isinstance(dep, str) else str(dep) if dep else None)
    return {
        "departure": dep,
        "origin_country": oc,
        "origin_postal": op,
        "arrival_country": dr.get("arrival_country"),
        "destination_zip_code": dr.get("destination_zip_code"),
    }


def _write_lane_data_row(
    ws,
    row: int,
    lane_idx: int,
    origin_country: object | None,
    origin_postal: object | None,
    destination_country: object | None,
    destination_postal: object | None,
    service_type: str | None,
    measurement: str,
    row_vals: list[object | None],
    layout: list[dict],
    column_specs: list[tuple[str, str, str | None]],
) -> None:
    ws.cell(row=row, column=1, value=lane_idx)
    ws.cell(row=row, column=2, value=origin_country)
    ws.cell(row=row, column=3, value=origin_postal)
    ws.cell(row=row, column=4, value=destination_country)
    ws.cell(row=row, column=5, value=destination_postal)
    ws.cell(row=row, column=6, value=service_type)
    ws.cell(row=row, column=7, value=measurement)
    for li, item in enumerate(layout):
        if item["type"] == "amount":
            idx = item["amount_index"]
            ws.cell(
                row=row,
                column=_layout_excel_col(li),
                value=_value_for_spec_column(
                    column_specs, idx, measurement, row_vals
                ),
            )


def transform_json_to_xlsx(
    data: dict,
    out_path: Path,
    sheet_title: str = "Rate Card",
) -> Path:
    main_costs = data.get("MainCosts") or []

    segments: list[dict] = []
    for block in main_costs:
        tab = str(block.get("tab_name") or "Tab")
        rateids = _segment_rate_ids(block)
        rateid = rateids[0] if rateids else None
        if not rateid and isinstance(block.get("rateid"), str):
            rid = block.get("rateid", "").strip()
            rateid = rid or None
            if rid and not rateids:
                rateids = [rid.lower()]
        rcm = block.get("rate_column_map") or []
        drows = block.get("rows") or []
        if not rcm or not drows:
            continue
        segments.append(
            {
                "tab": tab,
                "rateid": rateid,
                "rateids": rateids,
                "rcm": rcm,
                "rows": drows,
            }
        )

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    ws.title = sheet_title[:31]

    if not segments:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    main_column_specs = _build_main_column_specs(segments)
    main_spec_count = len(main_column_specs)
    column_specs = _append_global_targofix3_section(main_column_specs)
    spec_offsets = _spec_offsets_for_segments(segments, main_column_specs)
    total_cost_cols = len(column_specs)
    layout = _build_display_layout(column_specs)
    last_data_col = _layout_excel_col(len(layout) - 1) if layout else BASE_COLS

    merged_lanes: dict[tuple[str, str, str], dict] = {}
    lane_order: list[tuple[str, str, str]] = []

    for s in segments:
        for dr in s["rows"]:
            if _is_noise_tariff_row(dr):
                continue
            k = _lane_key(dr)
            if k not in merged_lanes:
                merged_lanes[k] = _shipment_shard(dr)
                lane_order.append(k)

    amounts: dict[tuple[str, str, str], list[object | None]] = {
        k: [None] * total_cost_cols for k in merged_lanes
    }

    _fill_amounts_from_segments(amounts, segments, column_specs, spec_offsets)

    for row_amt in amounts.values():
        _normalize_request_values_in_row(row_amt, column_specs)

    for row_amt in amounts.values():
        _fill_global_section_columns(row_amt, column_specs, main_spec_count)

    _strip_pallet_amounts_for_asterisk_lanes(amounts, column_specs)

    row = 1

    if _any_tab_needs_grouped_row(column_specs):
        _write_grouped_cost_row(ws, row, column_specs, layout)
        row += 1

    _write_merged_cost_name_row(ws, row, column_specs, layout)
    row += 1

    _write_merged_applies_if_row(ws, row, column_specs, layout, segments, spec_offsets)
    row += 1

    for li, item in enumerate(layout):
        if item["type"] != "amount":
            continue
        spec_idx = item["amount_index"]
        spec = column_specs[spec_idx]
        _tab, kind, _rid = spec[0], spec[1], spec[2]
        meta = _spec_meta(spec)
        if kind in EXTRA_COST_KINDS:
            continue
        if meta.get("pad_band"):
            rate_val = _rate_by_lane(KIND_PALLET)
        else:
            entry = _rcm_entry_for_spec(
                spec_idx, column_specs, segments, spec_offsets
            )
            if entry is None:
                continue
            rate_val = _rate_by_lane(_classify_cost_column(entry))
        ws.cell(row=row, column=_layout_excel_col(li), value=rate_val)
    row += 1

    for li, item in enumerate(layout):
        if item["type"] != "amount":
            continue
        spec_idx = item["amount_index"]
        spec = column_specs[spec_idx]
        _tab, kind, _rid = spec[0], spec[1], spec[2]
        meta = _spec_meta(spec)
        if kind in EXTRA_COST_KINDS:
            continue
        pad = meta.get("pad_band")
        if pad:
            band_val = pad
        else:
            entry = _rcm_entry_for_spec(
                spec_idx, column_specs, segments, spec_offsets
            )
            if entry is None:
                continue
            band_val = _band_label(entry, _classify_cost_column(entry))
        ws.cell(row=row, column=_layout_excel_col(li), value=band_val)
    row += 1

    for c, h in enumerate(SHIPMENT_HEADERS, start=1):
        ws.cell(row=row, column=c, value=h)
    _write_flat_currency_row(ws, row, layout, column_specs)
    row += 1

    lane_idx = 0
    for k in lane_order:
        shard = merged_lanes[k]
        oc = shard.get("origin_country")
        op = shard.get("origin_postal")
        dest_c = shard.get("arrival_country")
        dest_pc = _format_destination_postal(shard.get("destination_zip_code"))
        svc = _service_type_cell_value(
            k,
            segments,
            amounts[k],
            main_spec_count,
            spec_offsets,
            shard,
        )
        lane_idx += 1
        _write_lane_data_row(
            ws,
            row,
            lane_idx,
            oc,
            op,
            dest_c,
            dest_pc,
            svc,
            "Standard",
            amounts[k],
            layout,
            column_specs,
        )
        row += 1

    for k in lane_order:
        shard = merged_lanes[k]
        oc = shard.get("origin_country")
        op = shard.get("origin_postal")
        dest_c = shard.get("arrival_country")
        dest_pc = _format_destination_postal(shard.get("destination_zip_code"))
        svc = _service_type_cell_value(
            k,
            segments,
            amounts[k],
            main_spec_count,
            spec_offsets,
            shard,
            include_zalando=False,
        )
        lane_idx += 1
        _write_lane_data_row(
            ws,
            row,
            lane_idx,
            dest_c,
            dest_pc,
            oc,
            op,
            svc,
            "Return",
            amounts[k],
            layout,
            column_specs,
        )
        row += 1

    max_col = max(BASE_COLS + 5, last_data_col)
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _pick_json_interactive(folder: Path) -> Path:
    files = sorted(folder.glob("*.json"), key=lambda p: p.name.lower())
    if not files:
        raise SystemExit(f"No .json files in {folder}")
    print("Select JSON file:\n")
    for i, p in enumerate(files, 1):
        print(f"  {i}. {p.name}")
    while True:
        raw = input(f"Enter number (1–{len(files)}): ").strip()
        if raw.isdigit():
            n = int(raw)
            if 1 <= n <= len(files):
                return files[n - 1]
        print("Invalid choice.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert main.py tariff JSON to Excel rate card layout.",
    )
    parser.add_argument(
        "input_json",
        nargs="?",
        default=None,
        help="Path to JSON (default: pick from output/ interactively).",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output .xlsx path (default: same stem as JSON under output/).",
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"Folder for interactive JSON pick (default: {DEFAULT_INPUT_DIR}).",
    )
    args = parser.parse_args()

    if args.input_json:
        in_path = Path(args.input_json)
        if not in_path.is_absolute() and not in_path.exists():
            cand = args.input_dir / in_path
            if cand.is_file():
                in_path = cand
    else:
        in_path = _pick_json_interactive(args.input_dir)

    if not in_path.is_file():
        raise SystemExit(f"Input not found: {in_path}")

    if args.output:
        out_path = Path(args.output)
    else:
        out_path = in_path.with_suffix(".rate_card.xlsx")

    with open(in_path, encoding="utf-8") as f:
        data = json.load(f)

    transform_json_to_xlsx(data, out_path)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
