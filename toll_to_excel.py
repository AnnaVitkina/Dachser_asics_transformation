"""
Export structured toll matrices from main_toll.py JSON to Excel.

Source: toll tabs live in the Europe transport tariff workbook (e.g.
``Copy of Tariffs Europe for DCS - 2024-2027 - 20240401 - V14 (2).xlsx``),
not the appendix. Run ``main_toll.py`` on that file first; then point this
script at the resulting ``*toll*.json`` in ``output/`` (typically ``*.toll.json``).

All structured toll matrices on one worksheet, stacked by cost table.
WeightMatrixSegment 1+2 (and duplicate CostTable entries with the same name)
are merged into a single matrix. Layout per block:
  - Column A: ``Destination/Origin postal code zone equals Toll (DE) Zone 1``;
    country regions use ``Destination/Origin country region equals Toll (DE) Zone 1``
  - Reference catalogs: ``*.toll_postal_zones.txt`` and ``*.toll_country_regions.txt``
  - Columns B+: weight bands reformatted (From 1 to 50 kg -> <=50)
  - Cells: tariff amounts

Unstructured percentage surcharges are exported on a second worksheet (``Toll %``)
with columns ``% over costs`` and ``applies if``.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "output"

DEST_ZONE_EQUALS = "Destination postal code zone equals "
ORIGIN_ZONE_EQUALS = "Origin postal code zone equals "
DEST_COUNTRY_REGION_EQUALS = "Destination country region equals "
ORIGIN_COUNTRY_REGION_EQUALS = "Origin country region equals "

_COUNTRY_REGION_ZONING_FORMATS = frozenset(
    {
        "country_columns",
        "transit_destination_zones",
        "transit_concerned_countries",
    }
)

COMBINED_SHEET_TITLE = "Toll"
UNSTRUCTURED_SHEET_TITLE = "Toll %"
UNSTRUCTURED_HEADERS = ("% over costs", "applies if")

# Country name (normalized key) -> ISO 3166-1 alpha-2
_COUNTRY_NAME_TO_ISO: dict[str, str] = {
    "albania": "AL",
    "armenia": "AM",
    "austria": "AT",
    "azerbaijan": "AZ",
    "germany": "DE",
    "belarus": "BY",
    "belgium": "BE",
    "bosniahercegovina": "BA",
    "bosniaandherzegovina": "BA",
    "bulgaria": "BG",
    "croatia": "HR",
    "czechrepublic": "CZ",
    "denmark": "DK",
    "finland": "FI",
    "kosovo": "XK",
    "montenegro": "ME",
    "norway": "NO",
    "sweden": "SE",
    "estonia": "EE",
    "georgia": "GE",
    "greatbritain": "GB",
    "hungary": "HU",
    "iran": "IR",
    "iraq": "IQ",
    "ireland": "IE",
    "jordan": "JO",
    "kazakhstan": "KZ",
    "kyrgyzstan": "KG",
    "kyryzstan": "KG",
    "latvia": "LV",
    "lithuania": "LT",
    "macedonia": "MK",
    "moldavia": "MD",
    "moldova": "MD",
    "poland": "PL",
    "romania": "RO",
    "russia": "RU",
    "serbia": "RS",
    "slovakia": "SK",
    "syria": "SY",
    "tajikistan": "TJ",
    "turkey": "TR",
    "turkmenistan": "TM",
    "ukraine": "UA",
    "uzbekistan": "UZ",
    "unitedkingdom": "GB",
}

# Tokens in section titles -> hub country code(s)
_SECTION_TITLE_HUB: list[tuple[str, str | list[str]]] = [
    ("great britain and ireland", ["GB", "IE"]),
    ("great britain", "GB"),
    ("ireland", "IE"),
    ("poland", "PL"),
    ("hungary", "HU"),
    ("belgium", "BE"),
    ("denmark", "DK"),
    ("slovakia", "SK"),
    ("czech", "CZ"),
    ("germany", "DE"),
    ("austria", "AT"),
    ("france", "FR"),
]


def _discover_toll_json_files(directory: Path) -> list[Path]:
    """JSON files with 'toll' in the name (main_toll.py output in output/)."""
    files = [
        p
        for p in directory.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".json"
        and "toll" in p.name.lower()
    ]
    files.sort(
        key=lambda p: (
            0 if p.name.lower().endswith(".toll.json") else 1,
            p.name.lower(),
        )
    )
    return files


def _toll_matrices_output_path(toll_json: Path) -> Path:
    low = toll_json.name.lower()
    if low.endswith(".toll.json"):
        return toll_json.with_name(toll_json.name[: -len(".toll.json")] + ".toll_matrices.xlsx")
    return toll_json.with_suffix(".toll_matrices.xlsx")


def _toll_postal_zones_txt_path(toll_json: Path) -> Path:
    low = toll_json.name.lower()
    if low.endswith(".toll.json"):
        return toll_json.with_name(
            toll_json.name[: -len(".toll.json")] + ".toll_postal_zones.txt"
        )
    return toll_json.with_suffix(".toll_postal_zones.txt")


def _toll_country_regions_txt_path(toll_json: Path) -> Path:
    low = toll_json.name.lower()
    if low.endswith(".toll.json"):
        return toll_json.with_name(
            toll_json.name[: -len(".toll.json")] + ".toll_country_regions.txt"
        )
    return toll_json.with_suffix(".toll_country_regions.txt")

# Sheet tab name -> ISO-style country code for row labels
_TAB_COUNTRY_CODE: list[tuple[str, str]] = [
    ("germany", "DE"),
    ("austria", "AT"),
    ("czech", "CZ"),
    ("poland", "PL"),
    ("hungary", "HU"),
    ("slovakia", "SK"),
    ("belgium", "BE"),
    ("denmark", "DK"),
    ("france", "FR"),
    ("great britain", "GB"),
    ("ireland", "IE"),
]


def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def _country_code(tab_name: str, cost_name: str) -> str:
    blob = f"{tab_name} {cost_name}".lower()
    for token, code in _TAB_COUNTRY_CODE:
        if token in blob:
            return code
    return "XX"


def _zone_number(zone_name: str) -> str | None:
    z = _norm_ws(zone_name)
    m = re.match(r"(?i)zone\s*(\d+)\s*", z)
    if m:
        return m.group(1)
    return None


def _norm_zone_key(zone_name: str) -> str:
    return _norm_ws(zone_name).lower()


def _fallback_zone_label(tab_name: str, cost_name: str, zone_name: str) -> str:
    """When no Zoning row matches (e.g. distance-based DE groupage bands)."""
    cc = _country_code(tab_name, cost_name)
    zn = _zone_number(zone_name)
    if zn:
        return f"Toll ({cc}) Zone {zn}"
    low = _norm_ws(zone_name).lower()
    if low in ("all distances", "concerned countries"):
        return f"Toll ({cc}) {zone_name.strip()}"
    if zone_name.strip():
        return f"Toll ({cc}) {_norm_ws(zone_name)}"
    return f"Toll ({cc})"


def _is_country_region_zoning(zoning_entry: dict) -> bool:
    fmt = str(zoning_entry.get("ZoningFormat") or "").lower()
    if fmt in _COUNTRY_REGION_ZONING_FORMATS:
        return True
    if zoning_entry.get("Countries") and not zoning_entry.get("PostalCode"):
        return True
    return False


def _format_postal_codes(raw: str) -> str:
    """CP 07-32-33 -> 07,32,33; CP 66 -> 66."""
    s = _norm_ws(raw)
    s = re.sub(r"^(?:CP|CZ)\s*", "", s, flags=re.I)
    parts = [p for p in re.split(r"[^\d]+", s) if p]
    return ",".join(parts)


def _countries_iso_from_entry(entry: dict[str, Any]) -> list[str]:
    codes: list[str] = []
    seen: set[str] = set()
    for name in entry.get("Countries") or []:
        iso = _lookup_country_iso(str(name))
        if iso and iso not in seen:
            seen.add(iso)
            codes.append(iso)
    if not codes and entry.get("Country"):
        for part in str(entry["Country"]).split(","):
            iso = _lookup_country_iso(part.strip())
            if iso and iso not in seen:
                seen.add(iso)
                codes.append(iso)
    return codes


def _country_iso_from_postal_entry(
    entry: dict[str, Any],
    tab_name: str,
    cost_name: str,
) -> str:
    field = str(entry.get("Country") or "")
    if field:
        codes = _country_names_to_codes(field.split(",")[0].strip())
        if codes:
            return codes[0]
    return _country_code(tab_name, cost_name)


def _region_short_name(tab_name: str, cost_name: str, zone_name: str) -> str:
    cc = _country_code(tab_name, cost_name)
    zn = _zone_number(zone_name)
    if zn:
        return f"Toll ({cc}) Zone {zn}"
    label = _norm_ws(zone_name)
    return f"Toll ({cc}) {label}" if label else f"Toll ({cc})"


def _build_sheet_zoning_catalog(
    sheet: dict,
) -> tuple[
    dict[tuple[str, str], str],
    list[tuple[str, str, str]],
    list[tuple[str, str]],
]:
    """
    Per toll sheet: short names for apply-if, postal catalog rows, region catalog rows.
    Postal zone numbers are sequential in JSON Zoning order (Zone 6 : may be Zone 2).
    """
    tab = str(sheet.get("Tab") or "Toll")
    short_names: dict[tuple[str, str], str] = {}
    postal_rows: list[tuple[str, str, str]] = []
    region_rows: list[tuple[str, str]] = []
    postal_seq: dict[str, int] = {}

    for entry in sheet.get("Zoning") or []:
        zn = str(entry.get("ZoneName") or "")
        zk = _norm_zone_key(zn)
        applicable = entry.get("CostNameApplicable") or [""]
        cost0 = str(applicable[0])

        if _is_country_region_zoning(entry):
            name = _region_short_name(tab, cost0, zn)
            codes = _countries_iso_from_entry(entry)
            region_rows.append((name, _format_code_list(codes)))
            for cost in applicable:
                short_names[(_norm_cost_name(str(cost)), zk)] = name
        elif entry.get("PostalCode"):
            cc = _country_code(tab, cost0)
            postal_seq[cc] = postal_seq.get(cc, 0) + 1
            name = f"Toll ({cc}) Zone {postal_seq[cc]}"
            row_cc = _country_iso_from_postal_entry(entry, tab, cost0)
            postal_rows.append(
                (name, row_cc, _format_postal_codes(str(entry["PostalCode"])))
            )
            for cost in applicable:
                short_names[(_norm_cost_name(str(cost)), zk)] = name

    return short_names, postal_rows, region_rows


def _build_all_zoning_catalogs(
    data: dict,
) -> tuple[
    dict[tuple[str, str], str],
    list[tuple[str, str, str]],
    list[tuple[str, str]],
]:
    short_names: dict[tuple[str, str], str] = {}
    postal_rows: list[tuple[str, str, str]] = []
    region_rows: list[tuple[str, str]] = []
    for sheet in data.get("TollSheets") or []:
        sn, pr, rr = _build_sheet_zoning_catalog(sheet)
        short_names.update(sn)
        postal_rows.extend(pr)
        region_rows.extend(rr)
    return short_names, postal_rows, region_rows


def _dedupe_by_first_col(rows: list[tuple]) -> list[tuple]:
    seen: set[Any] = set()
    out: list[tuple] = []
    for row in rows:
        if row[0] in seen:
            continue
        seen.add(row[0])
        out.append(row)
    return out


def write_postal_zones_catalog(path: Path, rows: list[tuple[str, str, str]]) -> None:
    lines = [
        "TOLL POSTAL CODE ZONES",
        "=" * 60,
        "Name = Toll (CC) Zone N (sequential per country in Zoning order).",
        "Country = ISO 2-letter code. Postal code = CP/CZ digits from tariff.",
        "",
        "Name\tCountry\tPostal code",
    ]
    for name, country, postal in _dedupe_by_first_col(rows):
        lines.append(f"{name}\t{country}\t{postal}")
    if not rows:
        lines.append("(none)")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_country_regions_catalog(path: Path, rows: list[tuple[str, str]]) -> None:
    lines = [
        "TOLL COUNTRY REGIONS",
        "=" * 60,
        "Name = Toll (CC) Zone N from ZONE N in tariff. Country = ISO codes in region.",
        "",
        "Name\tCountry",
    ]
    for name, countries in _dedupe_by_first_col(rows):
        lines.append(f"{name}\t{countries}")
    if not rows:
        lines.append("(none)")
    path.write_text("\n".join(lines), encoding="utf-8")


def _build_zoning_lookup(sheet: dict) -> dict[tuple[str, str], dict[str, Any]]:
    lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for entry in sheet.get("Zoning") or []:
        zn = str(entry.get("ZoneName") or "")
        zk = _norm_zone_key(zn)
        for applicable in entry.get("CostNameApplicable") or []:
            lookup[(_norm_cost_name(str(applicable)), zk)] = entry
    return lookup


def _find_zoning_entry(
    lookup: dict[tuple[str, str], dict[str, Any]],
    cost_name: str,
    zone_name: str,
) -> dict[str, Any] | None:
    cn = _norm_cost_name(cost_name)
    zk = _norm_zone_key(zone_name)
    if (cn, zk) in lookup:
        return lookup[(cn, zk)]
    for (lcn, lzk), entry in lookup.items():
        if lcn != cn or not lzk:
            continue
        if zk == lzk or zk.startswith(lzk):
            return entry
    return None


def _lookup_short_zone_name(
    short_names: dict[tuple[str, str], str],
    cost_name: str,
    zone_name: str,
) -> str | None:
    cn = _norm_cost_name(cost_name)
    zk = _norm_zone_key(zone_name)
    if (cn, zk) in short_names:
        return short_names[(cn, zk)]
    for (lcn, lzk), name in short_names.items():
        if lcn == cn and lzk and (zk == lzk or zk.startswith(lzk)):
            return name
    return None


def _resolve_zone_row(
    tab_name: str,
    cost_name: str,
    zone_name: str,
    zoning_lookup: dict[tuple[str, str], dict[str, Any]],
    short_names: dict[tuple[str, str], str],
) -> tuple[str, list[str]]:
    """Return (internal zone key, apply-if rows) using Toll (CC) Zone N labels."""
    name = _lookup_short_zone_name(short_names, cost_name, zone_name)
    entry = _find_zoning_entry(zoning_lookup, cost_name, zone_name)
    if name is not None:
        if entry is not None and _is_country_region_zoning(entry):
            return name, [
                f"{DEST_COUNTRY_REGION_EQUALS}{name}",
                f"{ORIGIN_COUNTRY_REGION_EQUALS}{name}",
            ]
        return name, [
            f"{DEST_ZONE_EQUALS}{name}",
            f"{ORIGIN_ZONE_EQUALS}{name}",
        ]

    label = _fallback_zone_label(tab_name, cost_name, zone_name)
    return label, [
        f"{DEST_ZONE_EQUALS}{label}",
        f"{ORIGIN_ZONE_EQUALS}{label}",
    ]


def _reform_weight_header(weight: str) -> str:
    """From 1 to 50 kg -> <=50; Full truck unchanged."""
    w = _norm_ws(weight)
    if not w:
        return ""
    low = w.lower()
    if low in ("full truck", "complet", "complete", "ftl"):
        return "Full truck" if low != "complet" else "Complet"
    m = re.search(r"from\s+(\d+)\s+to\s+(\d+)\s*kg", low)
    if m:
        return f"<={m.group(2)}"
    m = re.search(r"from\s+(\d+)\s+kg", low)
    if m:
        return f">{m.group(1)}"
    m = re.search(r"up to\s+(\d+)\s*kg", low)
    if m:
        return f"<={m.group(1)}"
    m = re.search(r"(\d+)\s*kg", low)
    if m:
        return f"<={m.group(1)}"
    return w


def _weight_sort_key(header: str) -> tuple[int, str]:
    h = header.strip()
    if h.lower() in ("full truck", "complet"):
        return (2_000_000, h)
    m = re.search(r"<=(\d+)", h)
    if m:
        return (int(m.group(1)), h)
    m = re.search(r">(\d+)", h)
    if m:
        return (1_000_000 + int(m.group(1)), h)
    return (999_999_999, h)


def _norm_cost_name(name: str) -> str:
    return _norm_ws(name).lower()


def _merge_cost_tables_for_sheet(
    sheet: dict,
) -> list[tuple[str, str, dict[str, Any]]]:
    """Merge duplicate CostName entries and all WeightMatrixSegment rates per tab."""
    tab_name = str(sheet.get("Tab") or "Toll")
    merged: dict[str, dict[str, Any]] = {}
    order: list[str] = []

    for ct in sheet.get("CostTables") or []:
        cost_name = str(ct.get("CostName") or "Toll")
        key = _norm_cost_name(cost_name)
        if key not in merged:
            merged[key] = {"CostName": cost_name, "Rates": []}
            order.append(key)
        merged[key]["Rates"].extend(ct.get("Rates") or [])

    return [(tab_name, merged[key]["CostName"], merged[key]) for key in order]


def _build_matrix(
    cost_table: dict,
    tab_name: str,
    zoning_lookup: dict[tuple[str, str], dict[str, Any]],
    short_names: dict[tuple[str, str], str],
) -> tuple[list[tuple[str, list[str]]], list[str], dict[tuple[str, str], Any]] | None:
    rates = cost_table.get("Rates") or []
    if not rates:
        return None

    col_set: dict[str, None] = {}
    zone_rows: list[tuple[str, list[str]]] = []
    zone_keys: list[str] = []
    cells: dict[tuple[str, str], Any] = {}
    cost_name = str(cost_table.get("CostName") or "")

    for rate in rates:
        zone_name = str(rate.get("ZoneName") or "")
        zone_key, apply_rows = _resolve_zone_row(
            tab_name, cost_name, zone_name, zoning_lookup, short_names
        )
        if zone_key not in zone_keys:
            zone_keys.append(zone_key)
            zone_rows.append((zone_key, apply_rows))
        for band in rate.get("Bands") or []:
            w_raw = str(band.get("Weight") or "")
            col = _reform_weight_header(w_raw)
            if not col:
                continue
            col_set[col] = None
            cells[(zone_key, col)] = band.get("Cost")

    if not col_set or not zone_rows:
        return None

    columns = sorted(col_set.keys(), key=_weight_sort_key)
    return zone_rows, columns, cells


class TollMatrixBlock:
    __slots__ = ("tab_name", "cost_name", "zone_rows", "columns", "cells")

    def __init__(
        self,
        tab_name: str,
        cost_name: str,
        zone_rows: list[tuple[str, list[str]]],
        columns: list[str],
        cells: dict[tuple[str, str], Any],
    ) -> None:
        self.tab_name = tab_name
        self.cost_name = cost_name
        self.zone_rows = zone_rows
        self.columns = columns
        self.cells = cells


def _write_combined_toll_sheet(ws: Any, blocks: list[TollMatrixBlock]) -> None:
    row = 1
    max_cols = 2

    for bi, block in enumerate(blocks):
        if bi > 0:
            row += 1
        ws.cell(row=row, column=1, value="Source tab")
        ws.cell(row=row, column=2, value=block.tab_name)
        row += 1
        ws.cell(row=row, column=1, value="Cost table")
        ws.cell(row=row, column=2, value=block.cost_name)
        row += 1
        for ci, col in enumerate(block.columns, start=2):
            ws.cell(row=row, column=ci, value=col)
        row += 1
        for zone_key, apply_rows in block.zone_rows:
            for condition in apply_rows:
                ws.cell(row=row, column=1, value=condition)
                for ci, col in enumerate(block.columns, start=2):
                    val = block.cells.get((zone_key, col))
                    if val is not None:
                        ws.cell(row=row, column=ci, value=val)
                row += 1
        max_cols = max(max_cols, len(block.columns) + 1)

    for col_idx in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def _normalize_country_key(name: str) -> str:
    return re.sub(r"[^a-z]", "", name.lower())


def _lookup_country_iso(name: str) -> str | None:
    key = _normalize_country_key(name)
    if not key:
        return None
    if key in _COUNTRY_NAME_TO_ISO:
        return _COUNTRY_NAME_TO_ISO[key]
    for alias, code in _COUNTRY_NAME_TO_ISO.items():
        if alias in key or key in alias:
            return code
    return None


def _country_names_to_codes(text: str) -> list[str]:
    codes: list[str] = []
    seen: set[str] = set()
    for part in re.split(r",| and ", text):
        part = _norm_ws(part)
        if not part:
            continue
        code = _lookup_country_iso(part)
        if code and code not in seen:
            seen.add(code)
            codes.append(code)
    return codes


def _format_code_list(codes: list[str]) -> str:
    return ", ".join(codes)


def _pct_over_costs(rate: float) -> str:
    value = rate * 100
    text = f"{value:.4f}".rstrip("0").rstrip(".")
    return text


def _hub_codes_from_section_title(section_title: str) -> list[str]:
    low = _norm_ws(section_title).lower()
    for token, code in _SECTION_TITLE_HUB:
        if token in low:
            if isinstance(code, list):
                return list(code)
            return [code]
    tail = re.sub(r"^(?:transit\s+)?toll\s+on\s+(?:the\s+)?", "", low, flags=re.I)
    tail = _norm_ws(tail)
    if tail:
        codes = _country_names_to_codes(tail)
        if codes:
            return codes
    return []


def _is_transit_section(section_title: str) -> bool:
    return "transit" in _norm_ws(section_title).lower()


_DOMESTIC_SURCHARGE = "__domestic_surcharge__"


def _parse_surcharge_line(line: str) -> tuple[str, float] | None:
    """Return (countries text or domestic marker, rate) or None if not a rate row."""
    line = _norm_ws(line)
    if not line or "|" not in line:
        return None
    left, right = line.rsplit("|", 1)
    try:
        rate = float(right.strip().replace(",", "."))
    except ValueError:
        return None
    left = left.strip()
    if "surcharge on the costs" in left.lower():
        return _DOMESTIC_SURCHARGE, rate
    m = re.match(r"concerned countries\s*:?\s*(.*)", left, re.I)
    countries_text = m.group(1).strip() if m else left
    if not countries_text:
        return None
    return countries_text, rate


def _domestic_apply_rows(hub_codes: list[str], rate: float) -> list[tuple[str, str]]:
    pct = _pct_over_costs(rate)
    rows: list[tuple[str, str]] = []
    for code in hub_codes:
        rows.append((pct, f"Origin country equals {code}"))
        rows.append((pct, f"Destination country equals {code}"))
    return rows


def _transit_apply_rows(
    hub_code: str,
    concerned_codes: list[str],
    rate: float,
) -> list[tuple[str, str]]:
    if not concerned_codes:
        return []
    pct = _pct_over_costs(rate)
    listed = _format_code_list(concerned_codes)
    return [
        (
            pct,
            f"Origin country equals {listed} and Destination country equals {hub_code}",
        ),
        (
            pct,
            f"Origin country equals {hub_code} and Destination country equals {listed}",
        ),
    ]


class UnstructuredTollSection:
    __slots__ = ("source_tab", "section_title", "rows")

    def __init__(
        self,
        source_tab: str,
        section_title: str,
        rows: list[tuple[str, str]],
    ) -> None:
        self.source_tab = source_tab
        self.section_title = section_title
        self.rows = rows


def _expand_unstructured_block(
    source_tab: str,
    section_title: str,
    block_rows: list[Any],
) -> UnstructuredTollSection | None:
    title = _norm_ws(section_title)
    hub_codes = _hub_codes_from_section_title(title)
    if not hub_codes:
        return None

    out_rows: list[tuple[str, str]] = []
    is_transit = _is_transit_section(title)

    for raw in block_rows:
        parsed = _parse_surcharge_line(str(raw))
        if parsed is None:
            continue
        countries_text, rate = parsed
        if countries_text == _DOMESTIC_SURCHARGE:
            out_rows.extend(_domestic_apply_rows(hub_codes, rate))
        elif is_transit:
            concerned = _country_names_to_codes(countries_text)
            out_rows.extend(_transit_apply_rows(hub_codes[0], concerned, rate))
        else:
            concerned = _country_names_to_codes(countries_text)
            if concerned:
                out_rows.extend(_transit_apply_rows(hub_codes[0], concerned, rate))
            else:
                out_rows.extend(_domestic_apply_rows(hub_codes, rate))

    if not out_rows:
        return None
    return UnstructuredTollSection(source_tab, title, out_rows)


def _build_unstructured_sections(data: dict) -> list[UnstructuredTollSection]:
    sections: list[UnstructuredTollSection] = []
    for sheet in data.get("TollSheets") or []:
        tab = str(sheet.get("Tab") or "")
        for ub in sheet.get("UnstructuredBlocks") or []:
            sec = _expand_unstructured_block(
                tab,
                str(ub.get("SectionTitle") or ""),
                ub.get("Rows") or [],
            )
            if sec is not None:
                sections.append(sec)
    return sections


def _write_unstructured_toll_sheet(
    ws: Any,
    sections: list[UnstructuredTollSection],
) -> int:
    row = 1
    data_rows = 0
    for si, section in enumerate(sections):
        if si > 0:
            row += 1
        ws.cell(row=row, column=1, value="Source tab")
        ws.cell(row=row, column=2, value=section.source_tab)
        row += 1
        ws.cell(row=row, column=1, value="Section")
        ws.cell(row=row, column=2, value=section.section_title)
        row += 1
        for ci, header in enumerate(UNSTRUCTURED_HEADERS, start=1):
            ws.cell(row=row, column=ci, value=header)
        row += 1
        for pct, apply_if in section.rows:
            ws.cell(row=row, column=1, value=pct)
            ws.cell(row=row, column=2, value=apply_if)
            row += 1
            data_rows += 1
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 100
    return data_rows


def transform_toll_json_to_workbook(
    data: dict,
) -> tuple[Workbook, int, int, list[tuple[str, str, str]], list[tuple[str, str]]]:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active worksheet")
    ws.title = COMBINED_SHEET_TITLE[:31]

    unstructured_sections = _build_unstructured_sections(data)
    short_names, postal_catalog, region_catalog = _build_all_zoning_catalogs(data)

    blocks: list[TollMatrixBlock] = []
    for sheet in data.get("TollSheets") or []:
        zoning_lookup = _build_zoning_lookup(sheet)
        for tab_name, cost_name, merged in _merge_cost_tables_for_sheet(sheet):
            built = _build_matrix(
                merged, tab_name, zoning_lookup, short_names
            )
            if built is None:
                continue
            zone_rows, columns, cells = built
            blocks.append(
                TollMatrixBlock(tab_name, cost_name, zone_rows, columns, cells)
            )

    if blocks:
        _write_combined_toll_sheet(ws, blocks)
    else:
        ws.cell(row=1, column=1, value="No structured toll matrices found in JSON.")

    n_unstructured = 0
    if unstructured_sections:
        ws_pct = wb.create_sheet(title=UNSTRUCTURED_SHEET_TITLE[:31])
        n_unstructured = _write_unstructured_toll_sheet(ws_pct, unstructured_sections)

    return wb, len(blocks), n_unstructured, postal_catalog, region_catalog


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Export toll JSON (from main_toll.py / Europe tariffs workbook) "
            "to matrix Excel."
        ),
    )
    ap.add_argument(
        "input_json",
        nargs="?",
        default=None,
        help="Path to a JSON file with 'toll' in the name (e.g. *.toll.json).",
    )
    ap.add_argument("-o", "--output", type=Path, default=None)
    ap.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Folder to search for *toll*.json (default: {DEFAULT_OUTPUT_DIR})",
    )
    args = ap.parse_args()

    if args.input_json:
        in_path = Path(args.input_json)
        if not in_path.is_absolute() and not in_path.exists():
            cand = args.input_dir / in_path
            if cand.is_file():
                in_path = cand
    else:
        files = _discover_toll_json_files(args.input_dir)
        if not files:
            raise SystemExit(
                f"No JSON with 'toll' in the name under {args.input_dir}. "
                "Run main_toll.py on the Europe tariffs .xlsx first."
            )
        if len(files) == 1:
            in_path = files[0]
            print(f"Using toll JSON: {in_path.name}")
        else:
            print(
                "Toll JSON files (from main_toll.py on the Europe tariffs workbook):"
            )
            for i, p in enumerate(files, 1):
                print(f"  {i}. {p.name}")
            raw = input("Enter number: ").strip()
            if not raw.isdigit() or not (1 <= int(raw) <= len(files)):
                raise SystemExit("Invalid choice")
            in_path = files[int(raw) - 1]

    if not in_path.is_file():
        raise SystemExit(f"Input not found: {in_path}")
    if "toll" not in in_path.name.lower():
        raise SystemExit(
            f"Expected a toll JSON file (name should contain 'toll'): {in_path}"
        )

    out_path = Path(args.output) if args.output else _toll_matrices_output_path(in_path)
    out_postal = _toll_postal_zones_txt_path(in_path)
    out_regions = _toll_country_regions_txt_path(in_path)

    with open(in_path, encoding="utf-8") as f:
        data = json.load(f)

    wb, n_tables, n_unstructured, postal_cat, region_cat = (
        transform_toll_json_to_workbook(data)
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    write_postal_zones_catalog(out_postal, postal_cat)
    write_country_regions_catalog(out_regions, region_cat)

    print(
        f"Wrote: {out_path} ({len(wb.sheetnames)} sheet(s), "
        f"{n_tables} structured table(s), {n_unstructured} % row(s))"
    )
    print(f"Wrote: {out_postal} ({len(_dedupe_by_first_col(postal_cat))} zone(s))")
    print(f"Wrote: {out_regions} ({len(_dedupe_by_first_col(region_cat))} region(s))")


if __name__ == "__main__":
    main()
