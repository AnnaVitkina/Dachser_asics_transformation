"""
Read tariff .xlsx files from ./input, let the user pick one, and export all
relevant sheets to ./output/<filename>.json under a MainCosts array.
Sheets whose names contain "Toll" or "Fuel" (case-insensitive) are skipped.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

INPUT_DIR = Path(__file__).resolve().parent / "input"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"

# Stop scanning a sheet after this many rows (tariff grids are bounded).
MAX_SHEET_ROWS = 100_000
# End data when this many consecutive rows have no values in the table slice.
EMPTY_STREAK_STOP = 40

# Product / service line labels printed on tariff tabs (lowercase in JSON).
RATE_IDS: tuple[str, ...] = ("targospeed", "targofix", "directload", "targoflex")


def _cell_str(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, float) and cell == int(cell):
        return str(int(cell))
    return str(cell).strip()


def _row_strings(row: tuple) -> list[str]:
    return [_cell_str(c) for c in row]


def _should_skip_sheet(name: str) -> bool:
    lower = name.lower()
    return "toll" in lower or "fuel" in lower


def _find_rate_id_in_text(text: str) -> str | None:
    low = text.lower()
    for token in RATE_IDS:
        if re.search(
            rf"(?i)(?<![a-z0-9]){re.escape(token)}(?![a-z0-9])",
            low,
        ):
            return token
    return None


def _extract_rate_ids(rows: list[tuple], header_idx: int | None) -> list[str]:
    """
    Scan the sheet for all known product lines (targospeed, targofix, directload, targoflex).
    Returns unique ids in row-major discovery order.
    """
    found: list[str] = []
    seen: set[str] = set()
    hi = header_idx if header_idx is not None else 0
    limit = min(len(rows), max(120, hi + 15))
    for ri in range(limit):
        row = rows[ri]
        for cell in row[:45]:
            t = _cell_str(cell)
            if not t:
                continue
            token = _find_rate_id_in_text(t)
            if token and token not in seen:
                seen.add(token)
                found.append(token)
    return found


def _find_header_row(rows: list[tuple]) -> int | None:
    """First row that looks like the tariff table header (Departure + Arrival)."""
    for idx, row in enumerate(rows):
        parts = _row_strings(row)
        has_dep = any(p.lower() == "departure" for p in parts)
        has_arr = any("arrival" in p.lower() for p in parts)
        if has_dep and has_arr:
            return idx
    return None


def _header_column_bounds(header_parts: list[str]) -> tuple[int, int]:
    nonempty = [i for i, v in enumerate(header_parts) if v]
    if not nonempty:
        return 0, len(header_parts) - 1
    return min(nonempty), max(nonempty)


def _unique_headers(raw: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in raw:
        key = h if h else "Column"
        n = seen.get(key, 0)
        seen[key] = n + 1
        out.append(key if n == 0 else f"{key}_{n + 1}")
    return out


def _classify_core_column(header: str) -> str | None:
    """Map Excel header to a canonical row field, or None if it is a rate/amount column."""
    h = header.strip()
    if not h:
        return None
    low = h.lower()
    if low == "departure":
        return "departure"
    if "arrival" in low and "country" in low:
        return "arrival_country"
    if "destination" in low and "zip" in low:
        return "destination_zip_code"
    if "document" in low and "request" in low:
        return "document_request"
    return None


def _amount_key(index_one_based: int) -> str:
    return f"amount_{index_one_based:02d}"


def _normalize_cell_value(val: object) -> object:
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        return None if s == "" else s
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val


# Typical Excel currency / rate display (Number with 2 decimal places).
AMOUNT_EXCEL_DECIMALS = 2


def _format_amount_like_excel(val: object, decimals: int = AMOUNT_EXCEL_DECIMALS) -> object:
    """
    Export amount cells like Excel on screen: 2 decimal places for fractional rates,
    whole numbers as int when the rounded value is integral. Text such as 'On request'
    is left unchanged.
    """
    if val is None:
        return None
    if isinstance(val, str):
        low = val.lower()
        if low in ("on request", "upon request"):
            return val
        try:
            num = float(val.replace(",", ".").replace(" ", ""))
        except ValueError:
            return val
        return _format_amount_like_excel(num, decimals)
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        r = round(val, decimals)
        if r == int(r):
            return int(r)
        return r
    return val


def _is_weight_band_edge_token(text: str) -> bool:
    """Lower bound of a kg band in headers, e.g. 29.1, 59.1 — not a pallet flat-rate label."""
    s = text.strip().replace(",", ".")
    return bool(re.match(r"^\d+\.1$", s))


def _excel_header_is_pallet_count_not_kg_cap(excel_header: str) -> bool:
    """
    Integer pallet column headers (1, 2, 3, …) vs kg caps that use the same row as 29, 59, 100.
    """
    s = str(excel_header).strip()
    if not re.fullmatch(r"\d+", s):
        return False
    n = int(s)
    if n in (29, 59, 100):
        return False
    return 1 <= n <= 40


def _column_context_above(
    rows: list[tuple],
    header_idx: int,
    abs_col: int,
    c0: int,
    lookback: int = 10,
    excel_header: str | None = None,
) -> list[str]:
    """
    Non-empty cell texts in the same column from rows directly above the header row
    (e.g. weight-from / to bands, pallet labels, flat-rate titles).

    If that column is empty across those rows (merged titles, blank spacer row above the
    header), scan upward then left within the lookback window for a shared row label
    (e.g. 'Flat Rate (# of pallets per shpt)' on Chartering, 'Flate rate (# shipment)' on
    Fixed Time). Used for every sheet the same way — no tab-specific logic.

    When a pallet column (excel_header 1..40 except 29/59/100) only picks up a weight-band
    edge (e.g. '59.1') from a misaligned row above, ignore it so the merged flat-rate label
    is found instead.
    """
    start = max(0, header_idx - lookback)
    out: list[str] = []
    for ri in range(start, header_idx):
        row = rows[ri]
        if abs_col >= len(row):
            continue
        t = _cell_str(row[abs_col])
        if t:
            out.append(t)

    if out and excel_header is not None:
        if _excel_header_is_pallet_count_not_kg_cap(excel_header) and all(
            _is_weight_band_edge_token(x) for x in out
        ):
            out = []

    if out:
        return out

    # Merged / row labels: walk up from the row above the header (skip blank spacer rows)
    for ri in range(header_idx - 1, max(-1, header_idx - 1 - lookback), -1):
        if ri < 0:
            break
        row = rows[ri]
        for c in range(abs_col, c0 - 1, -1):
            if c < 0 or c >= len(row):
                continue
            t = _cell_str(row[c])
            if not t:
                continue
            if excel_header is not None and _excel_header_is_pallet_count_not_kg_cap(
                excel_header
            ) and _is_weight_band_edge_token(t):
                continue
            return [t]

    return []


def _build_column_plan(headers: list[str]) -> tuple[list[dict], int]:
    """
    Each entry: {"role": "core"|"amount", "field": str|None, "excel_header": str}
    field is set for core columns; amount columns are in sheet order.
    """
    plan: list[dict] = []
    amount_idx = 0
    for excel_header in headers:
        core = _classify_core_column(excel_header)
        if core:
            plan.append({"role": "core", "field": core, "excel_header": excel_header})
        else:
            amount_idx += 1
            plan.append(
                {
                    "role": "amount",
                    "field": _amount_key(amount_idx),
                    "excel_header": excel_header,
                }
            )
    return plan, amount_idx


def _looks_like_rateinfo(text: str) -> bool:
    low = text.lower()
    return ("available" in low or "avalaible" in low) and ("from" in low or " au " in low)


def _looks_like_ratename(text: str) -> bool:
    low = text.lower()
    if _looks_like_rateinfo(text):
        return False
    # Subheader / band rows above the main title — not the product rate name.
    if any(
        x in low
        for x in (
            "flat rate",
            "flate rate",
            "per shipment",
            "weight from",
            "# of pallets",
            "pallets per",
        )
    ):
        return False
    if "tariff" in low:
        return True
    return any(k in low for k in ("zalando", "chartering", "ltl / ftl"))


def _extract_meta(rows: list[tuple], header_idx: int) -> tuple[str | None, str | None]:
    ratename: str | None = None
    rateinfo: str | None = None
    for row in rows[:header_idx]:
        for cell in row[:30]:
            if cell is None:
                continue
            t = _cell_str(cell)
            if not t or len(t) < 4:
                continue
            if _looks_like_rateinfo(t):
                rateinfo = t
                continue
            if _looks_like_ratename(t):
                if ratename is None or len(t) > len(ratename):
                    ratename = t
    if ratename is None:
        for row in rows[:header_idx]:
            c0 = row[0] if row else None
            t = _cell_str(c0)
            if len(t) > 8:
                ratename = t
                break
    return ratename, rateinfo


def _row_has_values(slice_vals: list) -> bool:
    for v in slice_vals:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return True
    return False


CORE_ROW_FIELDS = (
    "departure",
    "arrival_country",
    "destination_zip_code",
    "document_request",
)


def _empty_standard_row() -> dict[str, object | None]:
    return {k: None for k in CORE_ROW_FIELDS}


def _parse_sheet(sheet_name: str, rows: list[tuple]) -> dict:
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return {
            "tab_name": sheet_name,
            "ratename": None,
            "rateinfo": None,
            "rateid": None,
            "rateids": [],
            "rate_column_map": [],
            "amount_slot_count": 0,
            "rows": [],
            "note": "No Departure/Arrival header row detected; skipped table extraction.",
        }

    ratename, rateinfo = _extract_meta(rows, header_idx)
    rateids = _extract_rate_ids(rows, header_idx)
    rateid = rateids[0] if rateids else None
    header_row = rows[header_idx]
    header_parts = _row_strings(header_row)
    c0, c1 = _header_column_bounds(header_parts)
    headers_raw = header_parts[c0 : c1 + 1]
    headers = _unique_headers(headers_raw)
    column_plan, amount_count = _build_column_plan(headers)

    rate_column_map: list[dict[str, object]] = []
    for slice_i, entry in enumerate(column_plan):
        if entry["role"] != "amount":
            continue
        abs_col = c0 + slice_i
        eh = entry["excel_header"]
        ctx = _column_context_above(rows, header_idx, abs_col, c0, excel_header=eh)
        desc = " · ".join(ctx) + " | " + eh if ctx else eh
        rate_column_map.append(
            {
                "amount_key": entry["field"],
                "excel_header": eh,
                "header_context": ctx,
                "column_description": desc,
            }
        )

    data_rows: list[dict[str, object | None]] = []
    empty_run = 0
    for row in rows[header_idx + 1 :]:
        slice_vals = list(row[c0 : c1 + 1])
        if not _row_has_values(slice_vals):
            empty_run += 1
            if empty_run >= EMPTY_STREAK_STOP:
                break
            continue
        empty_run = 0
        record = _empty_standard_row()
        for entry, raw_val in zip(column_plan, slice_vals):
            val = _normalize_cell_value(raw_val)
            if entry["role"] == "core":
                field = entry["field"]
                assert field is not None
                if field == "destination_zip_code" and val is not None:
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                    record[field] = str(val).strip() if not isinstance(val, str) else val.strip()
                else:
                    record[field] = val
            else:
                record[entry["field"]] = _format_amount_like_excel(val)
        data_rows.append(record)

    return {
        "tab_name": sheet_name,
        "ratename": ratename,
        "rateinfo": rateinfo,
        "rateid": rateid,
        "rateids": rateids,
        "rate_column_map": rate_column_map,
        "amount_slot_count": amount_count,
        "rows": data_rows,
    }


def _pad_rows_amount_slots(blocks: list[dict], global_max: int) -> None:
    """Ensure every row has amount_01 .. amount_{global_max}; missing slots are None."""
    if global_max <= 0:
        return
    for block in blocks:
        for row in block.get("rows") or []:
            for i in range(1, global_max + 1):
                key = _amount_key(i)
                if key not in row:
                    row[key] = None


def _load_workbook_rows(path: Path) -> dict[str, list[tuple]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: dict[str, list[tuple]] = {}
    try:
        for ws in wb.worksheets:
            if _should_skip_sheet(ws.title):
                continue
            rows: list[tuple] = []
            for i, row in enumerate(
                ws.iter_rows(min_row=1, max_row=MAX_SHEET_ROWS, values_only=True),
                start=1,
            ):
                rows.append(tuple(row))
            out[ws.title] = rows
    finally:
        wb.close()
    return out


def _pick_input_file() -> Path:
    if not INPUT_DIR.is_dir():
        raise SystemExit(f"Input folder not found: {INPUT_DIR}")

    files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not files:
        raise SystemExit(f"No .xlsx files in {INPUT_DIR}")

    print("Available Excel files in input/:\n")
    for i, p in enumerate(files, start=1):
        print(f"  {i}. {p.name}")
    print()
    while True:
        raw = input("Enter the number of the file to transform: ").strip()
        if not raw.isdigit():
            print("Please enter a positive integer.")
            continue
        n = int(raw)
        if 1 <= n <= len(files):
            return files[n - 1]
        print(f"Choose between 1 and {len(files)}.")


def transform_workbook(path: Path) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stem = path.stem
    safe = re.sub(r'[<>:"/\\|?*]', "_", stem)
    out_path = OUTPUT_DIR / f"{safe}.json"

    sheets_rows = _load_workbook_rows(path)
    main_costs = [_parse_sheet(name, rows) for name, rows in sheets_rows.items()]

    global_amount_slots = max(
        (b.get("amount_slot_count") or 0) for b in main_costs
    )
    _pad_rows_amount_slots(main_costs, global_amount_slots)

    standard_schema = {
        "core_row_fields": list(CORE_ROW_FIELDS),
        "amount_slot_count": global_amount_slots,
        "amount_fields": [_amount_key(i) for i in range(1, global_amount_slots + 1)],
        "notes": (
            "Each row uses the same keys. Rate columns are positional: amount_01 is the "
            "first non-core column in that sheet, etc. Every tab uses the same rules: each "
            "rate_column_map entry has header_context (values stacked above that column, or "
            "a shared merged label found by scanning up/left) and column_description "
            "(context + excel header). Numeric amount_* values are rounded to "
            f"{AMOUNT_EXCEL_DECIMALS} decimal places to match typical Excel display."
        ),
    }

    payload = {
        "source_file": path.name,
        "standard_schema": standard_schema,
        "MainCosts": main_costs,
    }
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return out_path


def main() -> None:
    chosen = _pick_input_file()
    out = transform_workbook(chosen)
    print(f"\nWrote: {out}")


if __name__ == "__main__":
    main()
