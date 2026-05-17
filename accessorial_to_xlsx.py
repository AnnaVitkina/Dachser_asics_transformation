"""
Build RA accessorial rate rows from main_accessorial.py JSON.

Toll matrices: main_toll.py on the Europe tariffs .xlsx, then toll_to_excel.py
on the JSON with toll in the name (e.g. *.toll.json in output/).

Output:
  - .xlsx with columns: Original cost name, Cost Name as in RA, Rate By,
    Apply if, Price, Currency, MIN, MAX
  - .txt report of costs present in JSON but not in the RA mapping, plus
    manual-check notes for special-delivery fees.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "output"

HEADERS = (
    "Original cost name",
    "Cost Name as in RA",
    "Rate By",
    "Apply if",
    "Price",
    "Currency",
    "MIN",
    "MAX",
)

DEFAULT_CURRENCY = "EUR"
APPLY_ALL_SHIPMENTS = "Apply to all shipments"

SPECIAL_DELIVERY_PATTERNS = (
    "special delivery",
    "special delivery fee",
    "fix time delivery",
    "fixed time delivery",
    "fixed delivery time",
)


def _norm_key(s: str) -> str:
    s = s.lower().strip()
    s = s.replace("\u201c", '"').replace("\u201d", '"').replace("\ufffd", "")
    s = re.sub(r"\s+", " ", s)
    return s


def _format_price(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if 0 < abs(value) < 1 and value != int(value):
            pct = value * 100
            s = f"{pct:.6g}"
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            return f"{s}%"
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _match_ra_name(original_label: str) -> str | None:
    key = _norm_key(original_label)

    def has(*parts: str) -> bool:
        return all(p in key for p in parts)

    rules: list[tuple[str, list[str]]] = [
        ("Fuel Surcharge", ["fuel surcharge according to our agreement"]),
        ("Trade Advice Surcharge", ['booking-in for "retail" (ah)']),
        ("Booking Fee", ['booking-in "automatic" (ap)']),
        ("Import Customs Clearance", ["import customs clearance"]),
        ("Difficulty Surcharge", ["metropolitan area (france)"]),
        ("Tail Lift Surcharge", ["ltl/ftl (direct loads) in france"]),
        ("CO2 Fee", ["co2 germany"]),
        ("Wasted Collection Fee", ["unsucccessful pick-up attempt caused"]),
        ("Channel Fee", ["crossing great britain and ireland"]),
        ("Export Customs Clearance", ["customs export accompanying document"]),
        ("Document Fee", ["transit note (t1-document, t2-document"]),
        (
            "Additional Customs",
            ["customs export accompanying document - additional line"],
        ),
        ("Computer Processing Fee", ["sad computer processing export"]),
        ("Delta Fee", []),
        ("ENS Fee", []),
        ("Phone Delivery Booking Fee", ['booking-in "per telephone" (ap)']),
        ("Second Delivery Fee", ["second delivery attempt domestic"]),
        ("Second Delivery Fee", ["second delivery attempt international"]),
        (
            "European large cities surcharge",
            ["european large cities surcharge"],
        ),
        ("SENT", ["uit romania, sent poland, utn bulgaria"]),
    ]
    if key == "delta":
        return "Delta Fee"
    if key == "ens":
        return "ENS Fee"
    if "per additional line" in key and "customs export" in key:
        return "Additional Customs"

    for ra_name, patterns in rules:
        for pat in patterns:
            if pat in key:
                if ra_name == "Export Customs Clearance" and "additional line" in key:
                    continue
                return ra_name
    return None


def _is_special_delivery_label(label: str) -> bool:
    low = _norm_key(label)
    return any(p in low for p in SPECIAL_DELIVERY_PATTERNS)


class CostRow:
    __slots__ = (
        "original",
        "source",
        "block",
        "apply_to",
        "price",
        "rate_by_extra",
        "apply_if_extra",
        "minimum_charge",
        "maximum_charge",
    )

    def __init__(
        self,
        original: str,
        source: str,
        block: str = "",
        apply_to: str = "",
        price: Any = None,
        rate_by_extra: str = "",
        apply_if_extra: str = "",
        minimum_charge: Any = None,
        maximum_charge: Any = None,
    ) -> None:
        self.original = original.strip()
        self.source = source
        self.block = block
        self.apply_to = apply_to
        self.price = price
        self.rate_by_extra = rate_by_extra
        self.apply_if_extra = apply_if_extra
        self.minimum_charge = minimum_charge
        self.maximum_charge = maximum_charge


def _accessorial_rows(data: dict) -> list[CostRow]:
    out: list[CostRow] = []
    src = data.get("source_file") or "accessorial"
    blocks = data.get("AccessorialCostBlocks") or {}
    sides = []
    if isinstance(blocks, dict):
        for side in ("Left", "Right"):
            sides.extend(blocks.get(side) or [])
    elif isinstance(blocks, list):
        sides = blocks

    for block in sides:
        btitle = str(block.get("BlockTitle") or "")
        for item in block.get("Items") or []:
            name = str(item.get("CostName") or "").strip()
            if not name:
                continue
            apply_to = str(item.get("ApplyTo") or "").strip()
            price = item.get("CostPrice")
            rows_to_add = [
                CostRow(
                    original=name,
                    source=src,
                    block=btitle,
                    apply_to=apply_to,
                    price=price,
                    minimum_charge=item.get("MinimumCharge"),
                    maximum_charge=item.get("MaximumCharge"),
                )
            ]
            if _norm_key(name) == "customs export accompanying document":
                rows_to_add.append(
                    CostRow(
                        original="Customs export accompanying document - Additional line",
                        source=src,
                        block=btitle,
                        apply_to="per additional line",
                        price=5.5,
                    )
                )
            out.extend(rows_to_add)
        # Block-level fuel surcharge (no item row with price)
        if _norm_key(btitle).startswith("fuel surcharge"):
            out.append(
                CostRow(
                    original=btitle,
                    source=src,
                    block=btitle,
                    apply_to="",
                    price="",
                )
            )

    return out


def _parse_weight_kg_condition(apply_to: str) -> tuple[str, str] | None:
    """Return (Rate By, Apply if) when apply_to encodes a weight band."""
    low = _norm_key(apply_to)
    m = re.search(r"up to\s+(\d+)\s*kg", low)
    if m:
        return ("Weight/kg", f"Weight/kg <={m.group(1)}")
    m = re.search(r"from\s+(\d+)\s*kg", low)
    if m:
        return ("Weight/kg", f"Weight/kg >{m.group(1)}")
    m = re.search(r"from\s+(\d+)\s+to\s+(\d+)\s*kg", low)
    if m:
        return (
            "Weight/kg",
            f"Weight/kg >={m.group(1)} and Weight/kg <={m.group(2)}",
        )
    m = re.search(r"from\s+(\d+)\s+to\s+(\d+)", low)
    if m and "kg" in low:
        return (
            "Weight/kg",
            f"Weight/kg >={m.group(1)} and Weight/kg <={m.group(2)}",
        )
    return None


def _parse_ppl_from_original(original: str) -> tuple[str, str] | None:
    """PPL limit in cost name, e.g. (<= 8 PPL)."""
    m = re.search(r"<=\s*(\d+)\s*ppl", original, re.I)
    if not m:
        m = re.search(r"\(\s*<=\s*(\d+)\s*ppl\s*\)", original, re.I)
    if not m:
        return None
    limit = m.group(1)
    rate_by = "Pallet/Pallet (PPL)"
    parts = [f"Pallet/Pallet <={limit}"]
    low = _norm_key(original)
    if "domestic" in low:
        parts.append("Domestic Shipments")
    elif "international" in low:
        parts.append("International Shipments")
    return rate_by, "; ".join(parts)


def _is_generic_per_unit(apply_to: str) -> bool:
    low = _norm_key(apply_to)
    generic = (
        "per consignment",
        "per declaration",
        "per pod",
        "per invoice",
        "per process",
        "per collection",
        "per hour",
        "per half hour",
        "per parcel and per day",
        "per 100 kg and per day",
        "per delivery truck",
        "per customs",
        "per additional line",
    )
    return low in generic


def _derive_ra_fields(cr: CostRow) -> tuple[str, str, str, str]:
    """
    Universal rules for Rate By, Apply if, MIN, MAX (never use block/group title).
    """
    min_out = _format_price(cr.minimum_charge) if cr.minimum_charge is not None else ""
    max_out = _format_price(cr.maximum_charge) if cr.maximum_charge is not None else ""

    apply_raw = (cr.apply_to or "").strip()
    orig = cr.original

    ppl = _parse_ppl_from_original(orig)
    if ppl:
        return ppl[0], ppl[1], min_out, max_out

    wk = _parse_weight_kg_condition(apply_raw)
    if wk:
        return wk[0], wk[1], min_out, max_out

    if "per 100 kg" in _norm_key(apply_raw):
        return "Weight/kg", APPLY_ALL_SHIPMENTS, min_out, max_out

    if re.search(r"per\s+pallet", apply_raw, re.I):
        return "Pallet/Pallet", APPLY_ALL_SHIPMENTS, min_out, max_out

    low_apply = _norm_key(apply_raw)
    if "%" in apply_raw and ("tariff" in low_apply or "transport" in low_apply):
        rate_by = "% Over costs"
        apply_parts: list[str] = []
        if "germany" in _norm_key(orig) and "transit" not in _norm_key(orig):
            apply_parts.append("Germany")
        elif "transit" in _norm_key(orig) and "germany" in _norm_key(orig):
            apply_parts.append("Transit via Germany")
        if "co2" in _norm_key(orig):
            apply_parts.append("CO2 surcharge")
        apply_if = "; ".join(apply_parts) if apply_parts else APPLY_ALL_SHIPMENTS
        return rate_by, apply_if, min_out, max_out

    if "surcharge on the costs of transport" in low_apply:
        return "% Over costs", APPLY_ALL_SHIPMENTS, min_out, max_out

    if _is_generic_per_unit(apply_raw):
        return "Flat", APPLY_ALL_SHIPMENTS, min_out, max_out

    m = re.match(r"^(per\s+\w+(?:\s+\w+)*)\s+(.+)$", apply_raw, re.I)
    if m:
        qualifier = m.group(2).strip()
        return "Flat", qualifier, min_out, max_out

    if not apply_raw:
        return "Flat", APPLY_ALL_SHIPMENTS, min_out, max_out

    return "Flat", APPLY_ALL_SHIPMENTS, min_out, max_out


def _to_output_row(cr: CostRow, ra_name: str) -> dict[str, str]:
    rate_by, apply_if, min_out, max_out = _derive_ra_fields(cr)
    price = _format_price(cr.price)
    currency = ""
    if price:
        if "€" in price or re.search(r"^\d+([.,]\d+)?$", price.replace(",", ".")):
            currency = DEFAULT_CURRENCY
        elif price.endswith("%"):
            currency = ""
        elif isinstance(cr.price, (int, float)):
            currency = DEFAULT_CURRENCY
    return {
        "Original cost name": cr.original,
        "Cost Name as in RA": ra_name,
        "Rate By": rate_by,
        "Apply if": apply_if,
        "Price": price,
        "Currency": currency,
        "MIN": min_out,
        "MAX": max_out,
    }


def _collect_all_records(accessorial_path: Path) -> list[CostRow]:
    with open(accessorial_path, encoding="utf-8") as f:
        return _accessorial_rows(json.load(f))


def build_ra_table(
    accessorial_path: Path,
) -> tuple[list[dict[str, str]], list[CostRow], list[CostRow]]:
    all_rows = _collect_all_records(accessorial_path)
    included: list[dict[str, str]] = []
    matched_ids: set[int] = set()
    special: list[CostRow] = []

    for i, cr in enumerate(all_rows):
        if _is_special_delivery_label(cr.original) or _is_special_delivery_label(
            cr.block
        ):
            special.append(cr)
        ra = _match_ra_name(cr.original)
        if ra:
            included.append(_to_output_row(cr, ra))
            matched_ids.add(i)

    special_keys = {
        (_norm_key(cr.original), _norm_key(cr.block)) for cr in special
    }
    unmatched = [
        cr
        for i, cr in enumerate(all_rows)
        if i not in matched_ids
        and (_norm_key(cr.original), _norm_key(cr.block)) not in special_keys
    ]
    return included, unmatched, special


def write_report(
    path: Path,
    unmatched: list[CostRow],
    special: list[CostRow],
) -> None:
    lines: list[str] = []
    lines.append("RA ACCESSORIAL — UNMATCHED COSTS")
    lines.append("=" * 72)
    lines.append(
        "Costs found in JSON that are not in the RA mapping list. "
        "Review for future mapping.\n"
    )
    if not unmatched:
        lines.append("(none)\n")
    else:
        for cr in unmatched:
            lines.append(f"Original: {cr.original}")
            lines.append(f"  Source file: {cr.source}")
            if cr.block:
                lines.append(f"  Block/Tab: {cr.block}")
            if cr.apply_to:
                lines.append(f"  ApplyTo / Rate basis: {cr.apply_to}")
            if cr.price is not None and str(cr.price).strip():
                lines.append(f"  Price: {_format_price(cr.price)}")
            if cr.apply_if_extra:
                lines.append(f"  Extra: {cr.apply_if_extra}")
            lines.append("")

    lines.append("\nMANUAL CHECK — SPECIAL / FIX TIME DELIVERY")
    lines.append("=" * 72)
    lines.append(
        "The following relate to Special Delivery Fee, Special Delivery Fee (KG), "
        "or Fix time delivery — verify in RA manually.\n"
    )
    seen: set[str] = set()
    for cr in special:
        key = (cr.original, cr.block)
        if key in seen:
            continue
        seen.add(key)
        lines.append(f"Original: {cr.original}")
        if cr.block:
            lines.append(f"  Block/Tab: {cr.block}")
        if cr.apply_to:
            lines.append(f"  ApplyTo: {cr.apply_to}")
        if cr.price is not None and str(cr.price).strip():
            lines.append(f"  Price: {_format_price(cr.price)}")
        lines.append("")

    if not special:
        lines.append("(no special-delivery labels detected in JSON)\n")

    path.write_text("\n".join(lines), encoding="utf-8")


def write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active worksheet")
    ws.title = "RA Accessorial"
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    for ri, row in enumerate(rows, start=2):
        for c, h in enumerate(HEADERS, start=1):
            ws.cell(row=ri, column=c, value=row.get(h, ""))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Build RA accessorial Excel from accessorial JSON.",
    )
    ap.add_argument(
        "input_json",
        nargs="?",
        default=None,
        help="Path to .accessorial.json (default: pick from output/).",
    )
    ap.add_argument("-o", "--output", type=Path, default=None)
    ap.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Default folder for stem-based inputs (default: {DEFAULT_OUTPUT_DIR})",
    )
    args = ap.parse_args()

    if args.input_json:
        acc_path = Path(args.input_json)
        if not acc_path.is_absolute() and not acc_path.exists():
            cand = args.output_dir / acc_path
            if cand.is_file():
                acc_path = cand
    else:
        files = sorted(args.output_dir.glob("*.accessorial.json"))
        if not files:
            raise SystemExit(f"No *.accessorial.json in {args.output_dir}")
        for i, p in enumerate(files, 1):
            print(f"  {i}. {p.name}")
        raw = input("Enter number: ").strip()
        if not raw.isdigit() or not (1 <= int(raw) <= len(files)):
            raise SystemExit("Invalid choice")
        acc_path = files[int(raw) - 1]

    if not acc_path.is_file():
        raise SystemExit(f"Input not found: {acc_path}")

    stem = acc_path.name.replace(".accessorial.json", "")
    out_xlsx = args.output or args.output_dir / f"{stem}.ra_accessorial.xlsx"
    out_txt = out_xlsx.with_suffix(".unmatched.txt")

    included, unmatched, special = build_ra_table(acc_path)
    write_xlsx(out_xlsx, included)
    write_report(out_txt, unmatched, special)

    print(f"Wrote: {out_xlsx} ({len(included)} rows)")
    print(f"Wrote: {out_txt} ({len(unmatched)} unmatched, {len(special)} special-delivery)")


if __name__ == "__main__":
    main()
