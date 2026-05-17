"""Shared Excel presentation helpers for simple table sheets."""

from __future__ import annotations

from copy import copy
from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def apply_simple_table_layout(
    ws: Worksheet,
    header_row: int = 1,
    *,
    freeze: bool = True,
    default_width: float = 18,
) -> None:
    """Bold header row, optional freeze pane, default column widths."""
    bold = Font(bold=True)
    for cell in ws[header_row]:
        if cell.value is not None:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    if freeze and ws.max_row >= header_row:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    for col in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(col)
        if letter not in ws.column_dimensions or ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = default_width


def apply_accessorial_sheet_layout(ws: Worksheet) -> None:
    apply_simple_table_layout(ws, header_row=1, default_width=20)
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36


def apply_toll_structured_sheet_layout(ws: Worksheet) -> None:
    """Style block headers and weight-band header rows on the Toll matrix sheet."""
    bold = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1):
        a = row[0].value
        a_s = str(a).strip() if a is not None else ""
        if a_s in ("Source tab", "Cost table"):
            for cell in row:
                if cell.value is not None:
                    cell.font = bold
            continue
        if a_s and (
            a_s.startswith("Destination postal code zone equals")
            or a_s.startswith("Origin postal code zone equals")
            or a_s.startswith("Destination country region equals")
            or a_s.startswith("Origin country region equals")
        ):
            continue
        b_val = row[1].value if len(row) > 1 else None
        if a_s == "" and b_val is not None:
            b_s = str(b_val)
            if b_s.startswith("<=") or b_s.startswith(">") or b_s.lower() in (
                "full truck",
                "complet",
            ):
                for cell in row:
                    if cell.value is not None:
                        cell.font = bold
    for col in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 16
    ws.column_dimensions["A"].width = 72


def apply_toll_pct_sheet_layout(ws: Worksheet) -> None:
    bold = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1):
        a = row[0].value
        b = row[1].value if len(row) > 1 else None
        a_s = str(a).strip() if a is not None else ""
        if a_s in ("Source tab", "Section", "% over costs") or b == "applies if":
            for cell in row:
                if cell.value is not None:
                    cell.font = bold
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 100
    if ws.max_row >= 1:
        ws.freeze_panes = "A2"


def copy_worksheet(source: Worksheet, target_wb: Any, new_title: str) -> Worksheet:
    """Copy cell values and styles into a new sheet on target_wb."""
    from openpyxl import Workbook

    if not isinstance(target_wb, Workbook):
        raise TypeError("target_wb must be an openpyxl Workbook")
    title = new_title[:31]
    if title in target_wb.sheetnames:
        for n in range(2, 100):
            cand = f"{title[:28]}_{n}"[:31]
            if cand not in target_wb.sheetnames:
                title = cand
                break
    target = target_wb.create_sheet(title=title)
    for row in source.iter_rows():
        for cell in row:
            t = target.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                t.font = copy(cell.font)
                t.border = copy(cell.border)
                t.fill = copy(cell.fill)
                t.number_format = cell.number_format
                t.protection = copy(cell.protection)
                t.alignment = copy(cell.alignment)
    for col, dim in source.column_dimensions.items():
        if dim.width is not None:
            target.column_dimensions[col].width = dim.width
    for row, dim in source.row_dimensions.items():
        if dim.height is not None:
            target.row_dimensions[row].height = dim.height
    if source.freeze_panes:
        target.freeze_panes = source.freeze_panes
    for merged in list(source.merged_cells.ranges):
        target.merge_cells(str(merged))
    return target
